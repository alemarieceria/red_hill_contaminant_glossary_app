"""Contract tests for the small, non-authoritative workbook pair."""

from hashlib import sha256
from pathlib import Path
import unittest
from unittest.mock import patch

from openpyxl.utils.cell import coordinate_to_tuple

from contaminant_pipeline import io_excel
from contaminant_pipeline.config import (
    GLOSSARY_TABLE_NAME,
    GLOSSARY_WORKBOOK_TYPE,
    METADATA_TABLE_NAME,
    REFERENCES_WORKBOOK_TYPE,
    WORKBOOK_SCHEMA_VERSION,
    validate_release_id,
)
from contaminant_pipeline.paths import (
    INCOMING_GLOSSARY_WORKBOOK,
    INCOMING_REFERENCES_WORKBOOK,
)
from contaminant_pipeline.schemas import (
    FOOTNOTE_HEADER_MAP,
    GLOSSARY_HEADER_MAP,
    REFERENCE_HEADER_MAP,
)
from fixture_paths import (
    FIXTURE_WORKBOOK_DIR,
    SYNTHETIC_GLOSSARY_WORKBOOK,
    SYNTHETIC_REFERENCES_WORKBOOK,
)


def sheet_named(snapshot, name):
    return next(sheet for sheet in snapshot.sheets if sheet.name == name)


def cell_at(sheet, coordinate):
    return next(cell for cell in sheet.cells if cell.coordinate == coordinate)


def row_values(sheet, row_number):
    cells = {
        coordinate_to_tuple(cell.coordinate): cell.value for cell in sheet.cells
    }
    return tuple(
        cells.get((row_number, column_number))
        for column_number in range(1, sheet.max_column + 1)
    )


def sha256_digest(path: Path) -> str:
    return sha256(path.read_bytes()).hexdigest()


class SyntheticWorkbookFixtureTests(unittest.TestCase):
    def test_fixture_paths_are_separate_from_authoritative_inputs(self) -> None:
        self.assertEqual(
            SYNTHETIC_GLOSSARY_WORKBOOK.parent,
            FIXTURE_WORKBOOK_DIR,
        )
        self.assertEqual(
            SYNTHETIC_REFERENCES_WORKBOOK.parent,
            FIXTURE_WORKBOOK_DIR,
        )
        self.assertEqual(
            SYNTHETIC_GLOSSARY_WORKBOOK.name,
            INCOMING_GLOSSARY_WORKBOOK.name,
        )
        self.assertEqual(
            SYNTHETIC_REFERENCES_WORKBOOK.name,
            INCOMING_REFERENCES_WORKBOOK.name,
        )
        self.assertNotEqual(
            SYNTHETIC_GLOSSARY_WORKBOOK,
            INCOMING_GLOSSARY_WORKBOOK,
        )
        self.assertNotEqual(
            SYNTHETIC_REFERENCES_WORKBOOK,
            INCOMING_REFERENCES_WORKBOOK,
        )

    def test_glossary_fixture_matches_the_source_contract(self) -> None:
        snapshot = io_excel.read_workbook(SYNTHETIC_GLOSSARY_WORKBOOK)

        self.assertEqual(snapshot.warnings, ())
        self.assertEqual(
            [sheet.name for sheet in snapshot.sheets],
            ["Introduction", "Glossary", "Footnotes", "Metadata"],
        )

        glossary = sheet_named(snapshot, "Glossary")
        self.assertEqual(row_values(glossary, 1), tuple(GLOSSARY_HEADER_MAP))
        self.assertEqual(
            [(table.name, table.reference) for table in glossary.tables],
            [(GLOSSARY_TABLE_NAME, "A1:AX4")],
        )
        self.assertEqual(cell_at(glossary, "A2").value, "Synthetic Alpha")
        self.assertEqual(
            cell_at(glossary, "E2").value,
            "Mock Alpha | Alpha Example",
        )
        self.assertFalse(cell_at(glossary, "K2").value)
        self.assertEqual(cell_at(glossary, "U2").value, "NA")
        self.assertEqual(cell_at(glossary, "AX3").value, "!!!!")
        self.assertEqual(cell_at(glossary, "AE4").value, "N/A")

        footnotes = sheet_named(snapshot, "Footnotes")
        self.assertEqual(row_values(footnotes, 1), tuple(FOOTNOTE_HEADER_MAP))
        self.assertEqual(row_values(footnotes, 2), ("A", "Synthetic explanatory footnote."))
        self.assertEqual(
            row_values(footnotes, 3),
            ("D", "Synthetic pesticide-status footnote."),
        )

        metadata = sheet_named(snapshot, "Metadata")
        self.assertEqual(
            [(table.name, table.reference) for table in metadata.tables],
            [(METADATA_TABLE_NAME, "A1:B4")],
        )
        self.assertEqual(cell_at(metadata, "B2").value, WORKBOOK_SCHEMA_VERSION)
        self.assertEqual(cell_at(metadata, "B3").value, GLOSSARY_WORKBOOK_TYPE)
        validate_release_id(cell_at(metadata, "B4").value)

    def test_references_fixture_matches_and_resolves_to_glossary(self) -> None:
        glossary_snapshot = io_excel.read_workbook(SYNTHETIC_GLOSSARY_WORKBOOK)
        references_snapshot = io_excel.read_workbook(
            SYNTHETIC_REFERENCES_WORKBOOK
        )

        self.assertEqual(references_snapshot.warnings, ())
        self.assertEqual(
            [sheet.name for sheet in references_snapshot.sheets],
            ["Sheet1", "Metadata"],
        )
        references = sheet_named(references_snapshot, "Sheet1")
        self.assertEqual(row_values(references, 1), tuple(REFERENCE_HEADER_MAP))
        self.assertEqual(references.max_row, 5)
        self.assertTrue(
            cell_at(references, "C2").value.startswith("https://example.com/")
        )

        glossary = sheet_named(glossary_snapshot, "Glossary")
        glossary_names_to_legacy_ids = {
            row_values(glossary, row_number)[0]: row_values(
                glossary, row_number
            )[2]
            for row_number in range(2, glossary.max_row + 1)
        }
        reference_names = [
            row_values(references, row_number)[0]
            for row_number in range(2, references.max_row + 1)
        ]
        self.assertEqual(
            glossary_names_to_legacy_ids,
            {
                "Synthetic Alpha": 901,
                "Synthetic Beta": 902,
                "Synthetic Mixture": 903,
            },
        )
        self.assertTrue(
            all(name in glossary_names_to_legacy_ids for name in reference_names)
        )
        self.assertEqual(reference_names.count("Synthetic Alpha"), 2)

        metadata = sheet_named(references_snapshot, "Metadata")
        self.assertEqual(
            [(table.name, table.reference) for table in metadata.tables],
            [(METADATA_TABLE_NAME, "A1:B4")],
        )
        self.assertEqual(cell_at(metadata, "B2").value, WORKBOOK_SCHEMA_VERSION)
        self.assertEqual(cell_at(metadata, "B3").value, REFERENCES_WORKBOOK_TYPE)
        validate_release_id(cell_at(metadata, "B4").value)

    def test_reading_fixtures_is_immutable_and_avoids_incoming_workbooks(
        self,
    ) -> None:
        fixture_paths = (
            SYNTHETIC_GLOSSARY_WORKBOOK,
            SYNTHETIC_REFERENCES_WORKBOOK,
        )
        original_digests = {path: sha256_digest(path) for path in fixture_paths}
        opened_paths = []
        original_load_workbook = io_excel.load_workbook

        def recording_load_workbook(path, *args, **kwargs):
            opened_paths.append(Path(path).resolve())
            return original_load_workbook(path, *args, **kwargs)

        with patch.object(
            io_excel,
            "load_workbook",
            side_effect=recording_load_workbook,
        ):
            for path in fixture_paths:
                io_excel.read_workbook(path)

        self.assertEqual(
            {path.resolve() for path in fixture_paths},
            set(opened_paths),
        )
        self.assertNotIn(INCOMING_GLOSSARY_WORKBOOK.resolve(), opened_paths)
        self.assertNotIn(INCOMING_REFERENCES_WORKBOOK.resolve(), opened_paths)
        self.assertEqual(
            {path: sha256_digest(path) for path in fixture_paths},
            original_digests,
        )


if __name__ == "__main__":
    unittest.main()

