"""Tests for the Phase 3.3a supervisor review handoff."""

from dataclasses import FrozenInstanceError, replace
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import load_workbook

from contaminant_pipeline.config import GLOSSARY_WORKBOOK_FILENAME
from contaminant_pipeline.paths import (
    CONTAMINANT_REGISTRY_PATH,
    INCOMING_DIR,
    INCOMING_GLOSSARY_WORKBOOK,
    INCOMING_REFERENCES_WORKBOOK,
    PIPELINE_ROOT,
    REFERENCE_CROSSWALK_PATH,
    review_output_dir,
    supervisor_review_workbook_path,
)
from contaminant_pipeline.scientific_review import (
    ResolvedCleanupRecord,
    ReviewReconciliationStatus,
    ScientificReviewProposal,
    ScientificReviewStatus,
    ScientificReviewType,
    build_scientific_review_package,
    reconcile_scientific_review_items,
    write_supervisor_review_workbook,
)
from contaminant_pipeline.scientific_validation import inspect_scientific_fields
from contaminant_pipeline.validate import validate_identity_relationships

from test_scientific_validation import (
    _contract,
    _copy_synthetic_pair,
    _identity_result,
    _make_scientifically_valid,
    _set_glossary_values,
)


def _file_state(path: Path) -> tuple[bytes, int]:
    stat = path.stat()
    return path.read_bytes(), stat.st_mtime_ns


def _replace_snapshot_value(identities, coordinate: str, value: object):
    contract = identities.workbook_contract
    snapshot = contract.raw_pair.glossary_snapshot
    sheets = tuple(
        replace(
            sheet,
            cells=tuple(
                replace(cell, value=value)
                if cell.coordinate == coordinate
                else cell
                for cell in sheet.cells
            ),
        )
        if sheet.name == "Glossary"
        else sheet
        for sheet in snapshot.sheets
    )
    revised_snapshot = replace(snapshot, sheets=sheets)
    revised_pair = replace(contract.raw_pair, glossary_snapshot=revised_snapshot)
    return replace(
        identities,
        workbook_contract=replace(contract, raw_pair=revised_pair),
    )


class ScientificReviewTests(unittest.TestCase):
    def test_optional_empty_text_is_null_warning_but_whitespace_is_error(self) -> None:
        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            incoming = _copy_synthetic_pair(root)
            _make_scientifically_valid(incoming)
            _set_glossary_values(
                incoming / GLOSSARY_WORKBOOK_FILENAME,
                (
                    (3, "Sources", None),
                    (4, "Sources", " \u00a0"),
                ),
            )
            identities = _identity_result(incoming, root)
            identities = _replace_snapshot_value(identities, "AP2", "")
            inspection = inspect_scientific_fields(identities)
            by_id = {row.id_contaminant: row for row in inspection.contaminants}
            self.assertIsNone(
                by_id["RHC-901"].parsed_values["source_notes_sources"]
            )
            self.assertIsNone(
                by_id["RHC-902"].parsed_values["source_notes_sources"]
            )
            warnings = [
                finding
                for finding in inspection.findings
                if finding.code == "pending_source_notes_sources"
            ]
            self.assertEqual(
                {(finding.id_contaminant, finding.source_value) for finding in warnings},
                {("RHC-901", "''"), ("RHC-902", "None")},
            )
            whitespace = [
                finding
                for finding in inspection.findings
                if finding.code == "invalid_source_notes_sources"
            ]
            self.assertEqual(len(whitespace), 1)
            self.assertEqual(whitespace[0].id_contaminant, "RHC-903")

    def test_builds_frozen_review_rows_and_applies_supplied_evidence(self) -> None:
        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            incoming = _copy_synthetic_pair(root)
            _make_scientifically_valid(incoming)
            _set_glossary_values(
                incoming / GLOSSARY_WORKBOOK_FILENAME,
                (
                    (2, "CASRN", "64-17-4"),
                    (3, "InChIKey", None),
                    (4, "CASRN", "NA"),
                    (4, "InChIKey", "N/A"),
                    (4, "Sources", ""),
                ),
            )
            identities = _identity_result(incoming, root)
            proposal = ScientificReviewProposal(
                id_contaminant="RHC-901",
                canonical_field="id_casrn",
                proposed_value="64-17-5",
                source_system="PubChem",
                source_record_id="CID 702",
                evidence_url="https://pubchem.ncbi.nlm.nih.gov/compound/702",
                retrieval_date="2026-08-10",
                rationale="Corrects the invalid check digit for the synthetic row.",
            )
            cleanup = ResolvedCleanupRecord(
                id_contaminant="RHC-903",
                id_name="Synthetic Mixture",
                workbook="contaminant_glossary.xlsx",
                sheet="Glossary",
                cell="AP4",
                canonical_field="source_notes_sources",
                old_value="Synthetic narrative.\u00a0",
                new_value="Synthetic narrative.",
                correction_type="trailing_nonbreaking_space",
            )
            package = build_scientific_review_package(
                identities,
                proposals=(proposal,),
                resolved_cleanup=(cleanup,),
            )

            self.assertEqual(package.data_release_id, "20000115-r2")
            self.assertEqual(len(package.resolved_cleanup), 1)
            types = {item.review_type for item in package.review_items}
            self.assertTrue(
                {
                    ScientificReviewType.INVALID_SCIENTIFIC_VALUE,
                    ScientificReviewType.UNKNOWN_IDENTIFIER,
                    ScientificReviewType.UNVERIFIED_NOT_APPLICABLE,
                    ScientificReviewType.PENDING_SOURCE,
                }.issubset(types)
            )
            casrn = next(
                item
                for item in package.review_items
                if item.id_contaminant == "RHC-901"
                and item.canonical_field == "id_casrn"
            )
            self.assertEqual(casrn.cell, "F2")
            self.assertEqual(casrn.proposed_value, "64-17-5")
            self.assertEqual(casrn.status, ScientificReviewStatus.PROPOSED)
            with self.assertRaises(FrozenInstanceError):
                casrn.status = ScientificReviewStatus.RESOLVED

    def test_rejects_unreviewable_or_incomplete_proposals(self) -> None:
        with self.assertRaisesRegex(ValueError, "requires source_system"):
            ScientificReviewProposal(
                id_contaminant="RHC-901",
                canonical_field="id_casrn",
                proposed_value="64-17-5",
            )
        with self.assertRaisesRegex(ValueError, "requires proposed_value"):
            ScientificReviewProposal(
                id_contaminant="RHC-901",
                canonical_field="id_casrn",
                source_system="PubChem",
                source_record_id="CID 702",
                evidence_url="https://pubchem.ncbi.nlm.nih.gov/compound/702",
                status=ScientificReviewStatus.APPROVED_VALUE,
                reviewer="Reviewer",
                review_date="2026-08-10",
            )
        with self.assertRaisesRegex(ValueError, "evidence requires"):
            ScientificReviewProposal(
                id_contaminant="RHC-901",
                canonical_field="id_casrn",
                source_system="PubChem",
                status=ScientificReviewStatus.NEEDS_REVIEW,
            )
        with self.assertRaisesRegex(ValueError, "resolved is assigned"):
            ScientificReviewProposal(
                id_contaminant="RHC-901",
                canonical_field="id_casrn",
                status=ScientificReviewStatus.RESOLVED,
            )

    def test_reconciles_exact_resolved_and_superseded_findings(self) -> None:
        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            incoming = _copy_synthetic_pair(root)
            _make_scientifically_valid(incoming)
            _set_glossary_values(
                incoming / GLOSSARY_WORKBOOK_FILENAME,
                ((2, "CASRN", "64-17-4"),),
            )
            identities = _identity_result(incoming, root)
            package = build_scientific_review_package(identities)
            target = next(
                item
                for item in package.review_items
                if item.id_contaminant == "RHC-901"
                and item.canonical_field == "id_casrn"
            )
            inspection = inspect_scientific_fields(identities)
            exact = reconcile_scientific_review_items(
                (target,), inspection.findings
            )
            self.assertEqual(
                exact[0].status, ReviewReconciliationStatus.STILL_FAILING
            )
            resolved = reconcile_scientific_review_items((target,), ())
            self.assertEqual(
                resolved[0].status, ReviewReconciliationStatus.RESOLVED
            )
            replacement_finding = replace(
                next(
                    finding
                    for finding in inspection.findings
                    if finding.id_contaminant == "RHC-901"
                    and finding.canonical_field == "id_casrn"
                ),
                source_row=99,
            )
            superseded = reconcile_scientific_review_items(
                (target,), (replacement_finding,)
            )
            self.assertEqual(
                superseded[0].status, ReviewReconciliationStatus.SUPERSEDED
            )

    def test_writes_separate_review_workbook_without_formulas_or_source_changes(self) -> None:
        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            incoming = _copy_synthetic_pair(root)
            _make_scientifically_valid(incoming)
            identities = _identity_result(incoming, root)
            package = build_scientific_review_package(identities)
            protected = {
                path: _file_state(path)
                for path in (
                    incoming / GLOSSARY_WORKBOOK_FILENAME,
                    INCOMING_GLOSSARY_WORKBOOK,
                    INCOMING_REFERENCES_WORKBOOK,
                )
            }
            output = root / "review" / "supervisor_review_3_3.xlsx"
            result = write_supervisor_review_workbook(package, output)
            self.assertEqual(result, output)
            self.assertTrue(output.is_file())

            workbook = load_workbook(output, data_only=False)
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    [
                        "Instructions",
                        "Identifier Review",
                        "Pending Sources",
                        "Resolved Cleanup",
                    ],
                )
                for sheet in workbook.worksheets:
                    self.assertFalse(
                        any(
                            cell.data_type == "f"
                            for row in sheet.iter_rows()
                            for cell in row
                        )
                    )
                headers = [
                    cell.value for cell in workbook["Identifier Review"][1]
                ]
                instruction_rows = {
                    row[0].value: row[1].value
                    for row in workbook["Instructions"].iter_rows()
                }
                self.assertEqual(
                    instruction_rows["Column layout"],
                    "Generated source/finding columns come first; reviewer "
                    "decision and evidence columns follow with explicit names.",
                )
                self.assertFalse(
                    any(
                        term in str(value).casefold()
                        for value in instruction_rows.values()
                        for term in ("column colors", "blue columns", "yellow columns")
                    )
                )
                self.assertIn("existing_value", headers)
                self.assertIn("status", headers)
                self.assertIn("evidence_url", headers)
                self.assertGreater(workbook["Identifier Review"].max_row, 1)
                self.assertTrue(
                    all(sheet.sheet_state == "visible" for sheet in workbook.worksheets)
                )
                for sheet in workbook.worksheets:
                    self.assertFalse(sheet.tables)
                    self.assertIsNone(sheet.auto_filter.ref)
                    self.assertIsNone(sheet.freeze_panes)
                    self.assertFalse(sheet.data_validations.dataValidation)
                    self.assertFalse(sheet.column_dimensions)
                    self.assertFalse(sheet.row_dimensions)
                    self.assertTrue(
                        all(
                            cell.style_id == 0
                            for row in sheet.iter_rows()
                            for cell in row
                        )
                    )
            finally:
                workbook.close()

            self.assertEqual(
                {path: _file_state(path) for path in protected}, protected
            )

    def test_central_review_paths_are_release_scoped(self) -> None:
        self.assertEqual(
            review_output_dir("20260716"),
            PIPELINE_ROOT / "data" / "04_output" / "reviews" / "20260716",
        )
        self.assertEqual(
            supervisor_review_workbook_path("20260716"),
            review_output_dir("20260716") / "supervisor_review_3_3.xlsx",
        )


class AuthoritativeScientificReviewTests(unittest.TestCase):
    def test_builds_authoritative_review_queue_without_modifying_sources(self) -> None:
        protected_paths = (
            INCOMING_GLOSSARY_WORKBOOK,
            INCOMING_REFERENCES_WORKBOOK,
            CONTAMINANT_REGISTRY_PATH,
            REFERENCE_CROSSWALK_PATH,
            PIPELINE_ROOT / "pyproject.toml",
            PIPELINE_ROOT / "uv.lock",
        )
        protected_before = {path: _file_state(path) for path in protected_paths}
        with TemporaryDirectory() as temporary:
            identities = validate_identity_relationships(
                _contract(INCOMING_DIR, Path(temporary))
            )
            package = build_scientific_review_package(identities)
            counts = {}
            for item in package.review_items:
                counts[item.review_type] = counts.get(item.review_type, 0) + 1
            self.assertEqual(
                counts,
                {
                    ScientificReviewType.INVALID_SCIENTIFIC_VALUE: 6,
                    ScientificReviewType.PENDING_SOURCE: 6,
                    ScientificReviewType.UNKNOWN_IDENTIFIER: 32,
                    ScientificReviewType.UNVERIFIED_NOT_APPLICABLE: 19,
                },
            )
            scientific_errors = {
                item.id_contaminant
                for item in package.review_items
                if item.review_type
                is ScientificReviewType.INVALID_SCIENTIFIC_VALUE
            }
            self.assertEqual(
                scientific_errors,
                {"RHC-015", "RHC-038", "RHC-071", "RHC-086", "RHC-106", "RHC-107"},
            )
            pending_sources = {
                item.id_contaminant
                for item in package.review_items
                if item.review_type is ScientificReviewType.PENDING_SOURCE
            }
            self.assertEqual(
                pending_sources,
                {"RHC-132", "RHC-133", "RHC-134", "RHC-135", "RHC-136", "RHC-137"},
            )

        self.assertEqual(
            {path: _file_state(path) for path in protected_paths}, protected_before
        )


if __name__ == "__main__":
    unittest.main()
