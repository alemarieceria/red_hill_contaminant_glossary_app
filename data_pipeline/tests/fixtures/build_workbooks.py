"""Build the deterministic, non-authoritative workbook test fixtures."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from contaminant_pipeline.config import (
    FOOTNOTES_SHEET_NAME,
    GLOSSARY_SHEET_NAME,
    GLOSSARY_TABLE_NAME,
    GLOSSARY_WORKBOOK_FILENAME,
    GLOSSARY_WORKBOOK_TYPE,
    INTRODUCTION_SHEET_NAME,
    METADATA_SHEET_NAME,
    METADATA_TABLE_NAME,
    REFERENCES_SHEET_NAME,
    REFERENCES_WORKBOOK_FILENAME,
    REFERENCES_WORKBOOK_TYPE,
    WORKBOOK_SCHEMA_VERSION,
)
from contaminant_pipeline.schemas import (
    FOOTNOTE_HEADER_MAP,
    GLOSSARY_HEADER_MAP,
    REFERENCE_HEADER_MAP,
)


OUTPUT_DIR = Path(__file__).resolve().parent / "workbooks"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FILL = PatternFill("solid", fgColor="D9EAF7")
TABLE_STYLE = "TableStyleMedium2"


def _style_header(sheet, column_count: int) -> None:
    for cell in sheet[1][:column_count]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 34
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _add_table(sheet, name: str, reference: str) -> None:
    table = Table(displayName=name, ref=reference)
    table.tableStyleInfo = TableStyleInfo(
        name=TABLE_STYLE,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def _add_metadata(workbook: Workbook, workbook_type: str, revision: str) -> None:
    sheet = workbook.create_sheet(METADATA_SHEET_NAME)
    rows = (
        ("key", "value"),
        ("schema_version", WORKBOOK_SCHEMA_VERSION),
        ("workbook_type", workbook_type),
        ("workbook_revision", revision),
    )
    for row in rows:
        sheet.append(row)
    _style_header(sheet, 2)
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 28
    _add_table(sheet, METADATA_TABLE_NAME, "A1:B4")


def build_glossary_workbook(path: Path) -> None:
    workbook = Workbook()
    introduction = workbook.active
    introduction.title = INTRODUCTION_SHEET_NAME
    introduction.append(["SYNTHETIC TEST FIXTURE — NOT AUTHORITATIVE DATA"])
    introduction.append(
        ["Small deterministic workbook used by contaminant-pipeline tests only."]
    )
    introduction["A1"].fill = TITLE_FILL
    introduction["A1"].font = Font(bold=True, size=14, color="1F1F1F")
    introduction.column_dimensions["A"].width = 74

    glossary = workbook.create_sheet(GLOSSARY_SHEET_NAME)
    glossary.append(list(GLOSSARY_HEADER_MAP))
    rows = (
        (
            "Synthetic Alpha",
            "Alpha, Synthetic",
            901,
            "X2",
            "Mock Alpha | Alpha Example",
            "100-00-1 | 100-00-2",
            "AAAAAAAAAAAAAA-BBBBBBBBBB-C",
            "Aliphatic",
            "Alkane",
            "Linear",
            False,
            False,
            False,
            True,
            False,
            False,
            2,
            0,
            None,
            0,
            "NA",
            1,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            False,
            True,
            0.1,
            0.2,
            None,
            None,
            True,
            0.15,
            True,
            False,
            None,
            "Synthetic source A",
            "Test-only record.",
            "A",
            True,
            False,
            None,
            False,
            False,
            None,
        ),
        (
            "Synthetic Beta",
            "Beta, Synthetic",
            902,
            "Y6",
            None,
            None,
            "CCCCCCCCCCCCCC-DDDDDDDDDD-E",
            "Aromatic",
            "Benzene",
            None,
            True,
            True,
            None,
            False,
            True,
            "Contaminant",
            6,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            False,
            False,
            None,
            None,
            "Internal synthetic check",
            "6.5-8.5",
            False,
            None,
            False,
            False,
            "A",
            "Synthetic source B",
            None,
            "D",
            False,
            True,
            False,
            None,
            None,
            "!!!!",
        ),
        (
            "Synthetic Mixture",
            "Mixture, Synthetic",
            903,
            "N/A",
            "Example Blend | Mock Mixture",
            None,
            None,
            "Mixture",
            "Mixture of pure compounds",
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            "NA",
            "NA",
            "NA",
            "NA",
            "NA",
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            0,
            "N/A",
            "N/A",
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            "Synthetic source C",
            "Represents explicit not-applicable values.",
            "A, D",
            None,
            None,
            None,
            None,
            None,
            None,
        ),
    )
    expected_columns = len(GLOSSARY_HEADER_MAP)
    for row in rows:
        if len(row) != expected_columns:
            raise ValueError(
                f"glossary fixture row has {len(row)} values; "
                f"expected {expected_columns}"
            )
        glossary.append(row)
    _style_header(glossary, expected_columns)
    glossary.column_dimensions["A"].width = 24
    glossary.column_dimensions["B"].width = 24
    for column in ("E", "F", "G", "AP", "AQ"):
        glossary.column_dimensions[column].width = 24
    _add_table(glossary, GLOSSARY_TABLE_NAME, "A1:AX4")

    footnotes = workbook.create_sheet(FOOTNOTES_SHEET_NAME)
    footnotes.append(list(FOOTNOTE_HEADER_MAP))
    footnotes.append(["A", "Synthetic explanatory footnote."])
    footnotes.append(["D", "Synthetic pesticide-status footnote."])
    _style_header(footnotes, 2)
    footnotes.column_dimensions["A"].width = 12
    footnotes.column_dimensions["B"].width = 44

    _add_metadata(workbook, GLOSSARY_WORKBOOK_TYPE, "20000115")
    workbook.save(path)
    workbook.close()


def build_references_workbook(path: Path) -> None:
    workbook = Workbook()
    references = workbook.active
    references.title = REFERENCES_SHEET_NAME
    references.append(list(REFERENCE_HEADER_MAP))
    references.append(
        [
            "Synthetic Alpha",
            "Synthetic standard A",
            "https://example.com/references/alpha-a",
        ]
    )
    references.append(
        [
            "Synthetic Alpha",
            "Synthetic standard B",
            "https://example.com/references/alpha-b",
        ]
    )
    references.append(
        [
            "Synthetic Beta",
            "Synthetic standard C",
            "https://example.com/references/beta",
        ]
    )
    references.append(
        [
            "Synthetic Mixture",
            "Synthetic standard D",
            "https://example.com/references/mixture",
        ]
    )
    _style_header(references, len(REFERENCE_HEADER_MAP))
    references.column_dimensions["A"].width = 24
    references.column_dimensions["B"].width = 24
    references.column_dimensions["C"].width = 48

    _add_metadata(workbook, REFERENCES_WORKBOOK_TYPE, "20000115-r2")
    workbook.save(path)
    workbook.close()


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    build_glossary_workbook(OUTPUT_DIR / GLOSSARY_WORKBOOK_FILENAME)
    build_references_workbook(OUTPUT_DIR / REFERENCES_WORKBOOK_FILENAME)


if __name__ == "__main__":
    main()

