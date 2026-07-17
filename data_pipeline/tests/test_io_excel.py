from contextlib import contextmanager
from pathlib import Path
import unittest
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.worksheet.table import Table

from contaminant_pipeline.config import validate_release_id
from contaminant_pipeline.io_excel import ExcelReadError, read_workbook
from contaminant_pipeline.paths import (
    INCOMING_GLOSSARY_WORKBOOK,
    INCOMING_REFERENCES_WORKBOOK,
)


@contextmanager
def temporary_xlsx_path():
    path = Path(__file__).parent / f".temporary-{uuid4().hex}.xlsx"
    try:
        yield path
    finally:
        path.unlink(missing_ok=True)


def sheet_named(snapshot, name):
    return next(sheet for sheet in snapshot.sheets if sheet.name == name)


def cell_at(sheet, coordinate):
    return next(cell for cell in sheet.cells if cell.coordinate == coordinate)


class ExcelReaderTests(unittest.TestCase):
    def test_reads_structure_values_and_formula_definitions(self) -> None:
        with temporary_xlsx_path() as path:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Data"
            sheet.append(["name", "calculated"])
            sheet.append(["example", "=1+2"])
            sheet.add_table(Table(displayName="DataTable", ref="A1:B2"))
            workbook.save(path)
            workbook.close()
            original_bytes = path.read_bytes()

            snapshot = read_workbook(path)

            self.assertEqual(path.read_bytes(), original_bytes)
            self.assertEqual(snapshot.warnings, ())
            self.assertEqual([sheet.name for sheet in snapshot.sheets], ["Data"])
            data_sheet = snapshot.sheets[0]
            self.assertEqual(data_sheet.max_row, 2)
            self.assertEqual(data_sheet.max_column, 2)
            self.assertEqual(data_sheet.tables[0].name, "DataTable")
            self.assertEqual(data_sheet.tables[0].reference, "A1:B2")
            self.assertEqual(cell_at(data_sheet, "A2").value, "example")
            self.assertEqual(cell_at(data_sheet, "B2").formula, "=1+2")
            self.assertIsNone(cell_at(data_sheet, "B2").value)

    def test_reports_blank_comment_authors_without_stopping(self) -> None:
        with temporary_xlsx_path() as path:
            workbook = Workbook()
            sheet = workbook.active
            sheet["A1"] = "value"
            sheet["A1"].comment = Comment("review note", "   ")
            workbook.save(path)
            workbook.close()

            snapshot = read_workbook(path)

            self.assertEqual(len(snapshot.warnings), 1)
            warning = snapshot.warnings[0]
            self.assertEqual(warning.sheet_name, "Sheet")
            self.assertEqual(warning.coordinate, "A1")
            self.assertIn("author", warning.message)
            self.assertEqual(cell_at(snapshot.sheets[0], "A1").value, "value")

    def test_rejects_missing_unsupported_and_corrupt_inputs(self) -> None:
        missing = Path(__file__).parent / ".missing-workbook.xlsx"
        with self.assertRaises(ExcelReadError):
            read_workbook(missing)
        with self.assertRaises(ExcelReadError):
            read_workbook(Path(__file__))

        with temporary_xlsx_path() as path:
            path.write_bytes(b"not an Excel workbook")
            with self.assertRaises(ExcelReadError):
                read_workbook(path)

    def test_reads_real_workbooks_without_modifying_them(self) -> None:
        original_bytes = {
            path: path.read_bytes()
            for path in (
                INCOMING_GLOSSARY_WORKBOOK,
                INCOMING_REFERENCES_WORKBOOK,
            )
        }

        glossary = read_workbook(INCOMING_GLOSSARY_WORKBOOK)
        references = read_workbook(INCOMING_REFERENCES_WORKBOOK)

        for path, content in original_bytes.items():
            with self.subTest(path=path.name):
                self.assertEqual(path.read_bytes(), content)

        self.assertEqual(
            [sheet.name for sheet in glossary.sheets],
            ["Introduction", "Glossary", "Footnotes", "Metadata"],
        )
        glossary_sheet = sheet_named(glossary, "Glossary")
        self.assertIn("Table_1", {table.name for table in glossary_sheet.tables})
        formula_cells = [cell for cell in glossary_sheet.cells if cell.formula]
        self.assertGreater(len(formula_cells), 0)
        self.assertTrue(any(cell.value is not None for cell in formula_cells))
        glossary_metadata = sheet_named(glossary, "Metadata")
        self.assertEqual(
            [(table.name, table.reference) for table in glossary_metadata.tables],
            [("MetadataTable", "A1:B4")],
        )
        validate_release_id(cell_at(glossary_metadata, "B4").value)

        self.assertEqual(
            [sheet.name for sheet in references.sheets],
            ["Sheet1", "Metadata"],
        )
        references_metadata = sheet_named(references, "Metadata")
        self.assertEqual(
            [(table.name, table.reference) for table in references_metadata.tables],
            [("MetadataTable", "A1:B4")],
        )
        validate_release_id(cell_at(references_metadata, "B4").value)


if __name__ == "__main__":
    unittest.main()
