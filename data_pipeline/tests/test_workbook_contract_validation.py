from dataclasses import FrozenInstanceError, replace
from pathlib import Path
import shutil
from tempfile import TemporaryDirectory
import unittest
from unittest.mock import patch

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.worksheet.table import Table

from contaminant_pipeline.config import (
    FOOTNOTES_SHEET_NAME,
    GLOSSARY_SHEET_NAME,
    GLOSSARY_TABLE_NAME,
    GLOSSARY_WORKBOOK_FILENAME,
    METADATA_SHEET_NAME,
    REFERENCES_SHEET_NAME,
    REFERENCES_WORKBOOK_FILENAME,
)
from contaminant_pipeline.intake import (
    GitSourceState,
    IncomingContractError,
    inventory_incoming_pair,
    publish_or_reuse_intake,
    read_incoming_pair,
)
import contaminant_pipeline.intake as intake_module
from contaminant_pipeline.paths import (
    CONTAMINANT_REGISTRY_PATH,
    INCOMING_DIR,
    INCOMING_GLOSSARY_WORKBOOK,
    INCOMING_REFERENCES_WORKBOOK,
    MANIFEST_DIR,
    OUTPUT_DIR,
    PIPELINE_ROOT,
    PROCESSED_DIR,
    PUBLIC_DATA_DIR,
    RAW_SNAPSHOTS_DIR,
    REFERENCE_CROSSWALK_PATH,
)
from contaminant_pipeline.schemas import (
    FOOTNOTE_HEADER_MAP,
    GLOSSARY_HEADER_MAP,
    REFERENCE_HEADER_MAP,
)
from contaminant_pipeline.validate import (
    ValidatedWorkbookContract,
    WorkbookContractError,
    validate_workbook_contract,
)
import contaminant_pipeline.validate as validate_module
from fixture_paths import (
    SYNTHETIC_GLOSSARY_WORKBOOK,
    SYNTHETIC_REFERENCES_WORKBOOK,
)


GENERATED_ROOTS = (
    MANIFEST_DIR,
    RAW_SNAPSHOTS_DIR,
    PROCESSED_DIR,
    OUTPUT_DIR,
    PUBLIC_DATA_DIR,
)


def _file_state(path: Path) -> tuple[bytes, int]:
    return path.read_bytes(), path.stat().st_mtime_ns


def _tree_state() -> tuple[tuple[object, ...], ...]:
    state = []
    for root in GENERATED_ROOTS:
        state.append((root, root.exists()))
        if root.exists():
            for path in sorted(root.rglob("*")):
                state.append(
                    (
                        path,
                        path.is_dir(),
                        path.is_symlink(),
                        path.read_bytes() if path.is_file() else None,
                    )
                )
    return tuple(state)


def _copy_synthetic_pair(root: Path) -> Path:
    incoming = root / "incoming"
    incoming.mkdir()
    shutil.copyfile(
        SYNTHETIC_GLOSSARY_WORKBOOK,
        incoming / GLOSSARY_WORKBOOK_FILENAME,
    )
    shutil.copyfile(
        SYNTHETIC_REFERENCES_WORKBOOK,
        incoming / REFERENCES_WORKBOOK_FILENAME,
    )
    return incoming


def _edit_workbook(path: Path, edit) -> None:
    workbook = load_workbook(path)
    try:
        edit(workbook)
        workbook.save(path)
    finally:
        workbook.close()


def _publish(incoming: Path, root: Path):
    inventory = inventory_incoming_pair(read_incoming_pair(incoming))
    return publish_or_reuse_intake(
        inventory,
        root / "raw",
        root / "manifest",
        source_git=GitSourceState("unknown", None),
    )


def _headers(result: ValidatedWorkbookContract, role: str, sheet_name: str):
    workbook = (
        result.raw_inventory.glossary_inventory
        if role == "glossary"
        else result.raw_inventory.references_inventory
    )
    sheet = next(item for item in workbook.worksheets if item.name == sheet_name)
    if sheet.tables:
        return tuple(header.value for header in sheet.tables[0].headers)
    return tuple(header.value for header in sheet.headers)


class WorkbookContractValidationTests(unittest.TestCase):
    def test_validates_the_complete_synthetic_raw_pair_once_and_returns_frozen_data(
        self,
    ) -> None:
        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            incoming = _copy_synthetic_pair(root)
            publication = _publish(incoming, root)
            source_before = {
                path: _file_state(path) for path in incoming.iterdir() if path.is_file()
            }
            raw_before = {
                publication.raw_snapshot.glossary_path: _file_state(
                    publication.raw_snapshot.glossary_path
                ),
                publication.raw_snapshot.references_path: _file_state(
                    publication.raw_snapshot.references_path
                ),
            }

            with patch.object(
                intake_module,
                "read_workbook",
                wraps=intake_module.read_workbook,
            ) as reader:
                result = validate_workbook_contract(publication)

            self.assertEqual(result.data_release_id, "20000115-r2")
            self.assertEqual(result.schema_version, "1.0.0")
            self.assertIs(result.intake_publication, publication)
            self.assertEqual(reader.call_count, 2)
            self.assertEqual(
                [call.args[0].name for call in reader.call_args_list],
                [GLOSSARY_WORKBOOK_FILENAME, REFERENCES_WORKBOOK_FILENAME],
            )
            self.assertEqual(
                result.raw_pair.glossary_snapshot.path.parent,
                publication.raw_snapshot.snapshot_dir.resolve(),
            )
            self.assertEqual(
                result.raw_pair.references_snapshot.path.parent,
                publication.raw_snapshot.snapshot_dir.resolve(),
            )
            self.assertNotIn("00_incoming", str(result.raw_pair.glossary_snapshot.path))
            self.assertEqual(
                _headers(result, "glossary", GLOSSARY_SHEET_NAME),
                tuple(GLOSSARY_HEADER_MAP),
            )
            self.assertEqual(
                _headers(result, "glossary", FOOTNOTES_SHEET_NAME),
                tuple(FOOTNOTE_HEADER_MAP),
            )
            self.assertEqual(
                _headers(result, "references", REFERENCES_SHEET_NAME),
                tuple(REFERENCE_HEADER_MAP),
            )
            self.assertEqual(
                result.raw_pair.glossary_snapshot.warnings,
                publication.inventory.incoming_pair.glossary_snapshot.warnings,
            )
            self.assertEqual(
                {_path: _file_state(_path) for _path in source_before},
                source_before,
            )
            self.assertEqual(
                {_path: _file_state(_path) for _path in raw_before},
                raw_before,
            )
            with self.assertRaises(FrozenInstanceError):
                result.data_release_id = "changed"

    def test_validation_is_deterministic_and_independent_of_removed_incoming_files(
        self,
    ) -> None:
        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            incoming = _copy_synthetic_pair(root)
            publication = _publish(incoming, root)
            shutil.rmtree(incoming)

            first = validate_workbook_contract(publication)
            second = validate_workbook_contract(publication)

            self.assertEqual(first, second)

    def test_preserves_nonblocking_excel_reader_warnings(self) -> None:
        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            incoming = _copy_synthetic_pair(root)
            _edit_workbook(
                incoming / GLOSSARY_WORKBOOK_FILENAME,
                lambda workbook: setattr(
                    workbook[GLOSSARY_SHEET_NAME]["A2"],
                    "comment",
                    Comment("review note", "   "),
                ),
            )
            publication = _publish(incoming, root)

            result = validate_workbook_contract(publication)

            self.assertEqual(len(result.raw_pair.glossary_snapshot.warnings), 1)
            self.assertEqual(
                result.raw_pair.glossary_snapshot.warnings,
                publication.inventory.incoming_pair.glossary_snapshot.warnings,
            )

    def test_accepts_header_reordering_because_column_order_is_nonsemantic(self) -> None:
        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            incoming = _copy_synthetic_pair(root)

            def reorder(workbook):
                sheet = workbook[GLOSSARY_SHEET_NAME]
                sheet["A1"], sheet["B1"] = sheet["B1"].value, sheet["A1"].value

            _edit_workbook(incoming / GLOSSARY_WORKBOOK_FILENAME, reorder)
            publication = _publish(incoming, root)

            result = validate_workbook_contract(publication)

            actual = _headers(result, "glossary", GLOSSARY_SHEET_NAME)
            self.assertEqual(set(actual), set(GLOSSARY_HEADER_MAP))
            self.assertNotEqual(actual, tuple(GLOSSARY_HEADER_MAP))

    def test_rejects_missing_and_unexpected_glossary_headers(self) -> None:
        cases = (
            (
                "missing",
                lambda sheet: (
                    sheet.delete_cols(len(GLOSSARY_HEADER_MAP)),
                    setattr(
                        sheet.tables[GLOSSARY_TABLE_NAME],
                        "ref",
                        "A1:AW4",
                    ),
                ),
                "missing.*SafeWaters data",
            ),
            (
                "unexpected",
                lambda sheet: (
                    setattr(sheet.tables[GLOSSARY_TABLE_NAME], "ref", "A1:AY4"),
                    setattr(sheet["AY1"], "value", "Unapproved field"),
                ),
                "unexpected.*Unapproved field",
            ),
        )
        for name, edit, message in cases:
            with self.subTest(name=name), TemporaryDirectory() as temporary:
                root = Path(temporary)
                incoming = _copy_synthetic_pair(root)
                _edit_workbook(
                    incoming / GLOSSARY_WORKBOOK_FILENAME,
                    lambda workbook, operation=edit: operation(
                        workbook[GLOSSARY_SHEET_NAME]
                    ),
                )
                publication = _publish(incoming, root)

                with self.assertRaisesRegex(WorkbookContractError, message):
                    validate_workbook_contract(publication)

    def test_rejects_renamed_case_and_whitespace_header_variants(self) -> None:
        cases = (
            (GLOSSARY_WORKBOOK_FILENAME, GLOSSARY_SHEET_NAME, "A1", "Renamed"),
            (REFERENCES_WORKBOOK_FILENAME, REFERENCES_SHEET_NAME, "A1", "Compound_name"),
            (REFERENCES_WORKBOOK_FILENAME, REFERENCES_SHEET_NAME, "A1", " compound_name"),
            (GLOSSARY_WORKBOOK_FILENAME, FOOTNOTES_SHEET_NAME, "A1", "footnote_id"),
        )
        for filename, sheet_name, coordinate, replacement in cases:
            with self.subTest(replacement=replacement), TemporaryDirectory() as temporary:
                root = Path(temporary)
                incoming = _copy_synthetic_pair(root)
                _edit_workbook(
                    incoming / filename,
                    lambda workbook, value=replacement: setattr(
                        workbook[sheet_name][coordinate], "value", value
                    ),
                )
                publication = _publish(incoming, root)

                with self.assertRaisesRegex(
                    WorkbookContractError,
                    "headers do not match schema",
                ):
                    validate_workbook_contract(publication)

    def test_wraps_duplicate_blank_nontext_and_metadata_header_failures(self) -> None:
        cases = (
            (GLOSSARY_SHEET_NAME, "B1", tuple(GLOSSARY_HEADER_MAP)[0], "duplicate"),
            (GLOSSARY_SHEET_NAME, "B1", None, "nonblank literal text"),
            (GLOSSARY_SHEET_NAME, "B1", 7, "nonblank literal text"),
            (METADATA_SHEET_NAME, "A1", "Key", "headers must be exactly"),
        )
        for sheet_name, coordinate, value, detail in cases:
            with self.subTest(sheet=sheet_name, value=value), TemporaryDirectory() as temporary:
                root = Path(temporary)
                incoming = _copy_synthetic_pair(root)
                publication = _publish(incoming, root)
                raw_glossary = publication.raw_snapshot.glossary_path
                _edit_workbook(
                    raw_glossary,
                    lambda workbook, replacement=value: setattr(
                        workbook[sheet_name][coordinate], "value", replacement
                    ),
                )
                changed_before = _file_state(raw_glossary)

                with self.assertRaisesRegex(
                    WorkbookContractError,
                    f"raw workbook contract failed:.*{detail}",
                ) as caught:
                    validate_workbook_contract(publication)

                self.assertIsInstance(caught.exception.__cause__, IncomingContractError)
                self.assertEqual(_file_state(raw_glossary), changed_before)

    def test_rejects_wrong_input_and_inconsistent_completed_intake(self) -> None:
        with self.assertRaisesRegex(
            WorkbookContractError,
            "completed IntakePublication",
        ):
            validate_workbook_contract(object())

        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            publication = _publish(_copy_synthetic_pair(root), root)
            inconsistent_raw = replace(
                publication.raw_snapshot,
                data_release_id="20000116",
            )
            inconsistent_manifest = replace(
                publication.manifest,
                raw_snapshot=inconsistent_raw,
                manifest=replace(
                    publication.manifest.manifest,
                    data_release_id="20000116",
                ),
            )
            inconsistent = replace(
                publication,
                raw_snapshot=inconsistent_raw,
                manifest=inconsistent_manifest,
            )

            with self.assertRaisesRegex(
                WorkbookContractError,
                "identities do not agree",
            ):
                validate_workbook_contract(inconsistent)

            changed_inventory = replace(
                publication.inventory,
                glossary_inventory=replace(
                    publication.inventory.glossary_inventory,
                    warning_count=(
                        publication.inventory.glossary_inventory.warning_count + 1
                    ),
                ),
            )
            with self.assertRaisesRegex(
                WorkbookContractError,
                "does not retain its accepted inventory",
            ):
                validate_workbook_contract(
                    replace(publication, inventory=changed_inventory)
                )

    def test_wraps_missing_file_sheet_table_and_unsupported_schema_failures(self) -> None:
        def missing_file(publication):
            publication.raw_snapshot.references_path.unlink()

        def missing_sheet(publication):
            _edit_workbook(
                publication.raw_snapshot.glossary_path,
                lambda workbook: workbook.remove(workbook[FOOTNOTES_SHEET_NAME]),
            )

        def extra_sheet(publication):
            _edit_workbook(
                publication.raw_snapshot.glossary_path,
                lambda workbook: workbook.create_sheet("Unexpected"),
            )

        def missing_table(publication):
            _edit_workbook(
                publication.raw_snapshot.glossary_path,
                lambda workbook: workbook[GLOSSARY_SHEET_NAME].tables.clear(),
            )

        def extra_table(publication):
            def add_table(workbook):
                sheet = workbook["Introduction"]
                sheet["A20"] = "header"
                sheet["A21"] = "value"
                sheet.add_table(Table(displayName="UnexpectedTable", ref="A20:A21"))

            _edit_workbook(publication.raw_snapshot.glossary_path, add_table)

        def unsupported_schema(publication):
            def change_schema(workbook):
                sheet = workbook[METADATA_SHEET_NAME]
                schema_row = next(
                    row
                    for row in range(1, sheet.max_row + 1)
                    if sheet.cell(row, 1).value == "schema_version"
                )
                sheet.cell(schema_row, 2).value = "9.0.0"

            _edit_workbook(
                publication.raw_snapshot.glossary_path,
                change_schema,
            )

        cases = (
            (missing_file, "could not read required references"),
            (missing_sheet, "missing sheets"),
            (extra_sheet, "unknown sheets"),
            (missing_table, "expected tables"),
            (extra_table, "expected tables"),
            (unsupported_schema, "schema versions do not match|unsupported"),
        )
        for mutate, detail in cases:
            with self.subTest(detail=detail), TemporaryDirectory() as temporary:
                root = Path(temporary)
                publication = _publish(_copy_synthetic_pair(root), root)
                mutate(publication)

                with self.assertRaisesRegex(
                    WorkbookContractError,
                    f"raw workbook contract failed:.*({detail})",
                ) as caught:
                    validate_workbook_contract(publication)

                self.assertIsInstance(caught.exception.__cause__, IncomingContractError)

    def test_rejects_raw_bytes_changed_without_a_structural_error(self) -> None:
        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            publication = _publish(_copy_synthetic_pair(root), root)
            raw_glossary = publication.raw_snapshot.glossary_path
            _edit_workbook(
                raw_glossary,
                lambda workbook: setattr(
                    workbook[GLOSSARY_SHEET_NAME]["A2"],
                    "value",
                    "Changed fictional name",
                ),
            )
            changed_before = _file_state(raw_glossary)

            with self.assertRaisesRegex(
                WorkbookContractError,
                "raw snapshot (byte size|SHA-256) does not match",
            ):
                validate_workbook_contract(publication)

            self.assertEqual(_file_state(raw_glossary), changed_before)

    def test_fails_closed_without_an_explicit_supported_schema_header_contract(self) -> None:
        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            publication = _publish(_copy_synthetic_pair(root), root)

            with patch.object(validate_module, "_SCHEMA_HEADER_CONTRACTS", {}):
                with self.assertRaisesRegex(
                    WorkbookContractError,
                    "no workbook header contract",
                ):
                    validate_workbook_contract(publication)


class AuthoritativeWorkbookContractValidationTests(unittest.TestCase):
    def test_validates_current_release_from_disposable_raw_snapshots_only(self) -> None:
        protected_paths = (
            INCOMING_GLOSSARY_WORKBOOK,
            INCOMING_REFERENCES_WORKBOOK,
            CONTAMINANT_REGISTRY_PATH,
            REFERENCE_CROSSWALK_PATH,
            PIPELINE_ROOT / "pyproject.toml",
            PIPELINE_ROOT / "uv.lock",
        )
        protected_before = {path: _file_state(path) for path in protected_paths}
        generated_before = _tree_state()

        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            publication = publish_or_reuse_intake(
                inventory_incoming_pair(read_incoming_pair(INCOMING_DIR)),
                root / "raw",
                root / "manifest",
                source_git=GitSourceState("unknown", None),
            )
            raw_before = {
                publication.raw_snapshot.glossary_path: _file_state(
                    publication.raw_snapshot.glossary_path
                ),
                publication.raw_snapshot.references_path: _file_state(
                    publication.raw_snapshot.references_path
                ),
            }

            result = validate_workbook_contract(publication)

            self.assertEqual(result.data_release_id, "20260716")
            self.assertEqual(result.schema_version, "1.0.0")
            self.assertEqual(
                set(_headers(result, "glossary", GLOSSARY_SHEET_NAME)),
                set(GLOSSARY_HEADER_MAP),
            )
            self.assertEqual(
                set(_headers(result, "glossary", FOOTNOTES_SHEET_NAME)),
                set(FOOTNOTE_HEADER_MAP),
            )
            self.assertEqual(
                set(_headers(result, "references", REFERENCES_SHEET_NAME)),
                set(REFERENCE_HEADER_MAP),
            )
            self.assertEqual(
                {path: _file_state(path) for path in raw_before},
                raw_before,
            )

        self.assertEqual(
            {path: _file_state(path) for path in protected_paths},
            protected_before,
        )
        self.assertEqual(_tree_state(), generated_before)


if __name__ == "__main__":
    unittest.main()
