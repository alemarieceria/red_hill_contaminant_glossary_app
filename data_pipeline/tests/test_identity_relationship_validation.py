"""Tests for release-aware identity and relationship validation."""

from dataclasses import FrozenInstanceError
from pathlib import Path
import shutil
from tempfile import TemporaryDirectory
import unittest

from openpyxl import load_workbook

from contaminant_pipeline.config import (
    GLOSSARY_SHEET_NAME,
    GLOSSARY_WORKBOOK_FILENAME,
    REFERENCES_SHEET_NAME,
    REFERENCES_WORKBOOK_FILENAME,
    release_order_key,
)
from contaminant_pipeline.crosswalk import ReferenceResolutionMethod
from contaminant_pipeline.intake import (
    GitSourceState,
    inventory_incoming_pair,
    publish_or_reuse_intake,
    read_incoming_pair,
)
from contaminant_pipeline.paths import (
    CONTAMINANT_REGISTRY_PATH,
    INCOMING_DIR,
    INCOMING_GLOSSARY_WORKBOOK,
    INCOMING_REFERENCES_WORKBOOK,
    PIPELINE_ROOT,
    REFERENCE_CROSSWALK_PATH,
)
from contaminant_pipeline.registry_assets import (
    RegistryEntry,
    RegistryStatus,
    TrackedCrosswalkEntry,
    serialize_crosswalk,
    serialize_registry,
)
from contaminant_pipeline.validate import (
    IdentityRelationshipValidationError,
    ValidationSeverity,
    validate_identity_relationships,
    validate_workbook_contract,
)

from fixture_paths import (
    SYNTHETIC_GLOSSARY_WORKBOOK,
    SYNTHETIC_REFERENCES_WORKBOOK,
)


GENERATED_ROOTS = (
    PIPELINE_ROOT / "data" / "01_manifest",
    PIPELINE_ROOT / "data" / "02_raw_snapshots",
    PIPELINE_ROOT / "data" / "03_processed",
    PIPELINE_ROOT / "data" / "04_output",
    PIPELINE_ROOT.parent / "public" / "data",
)


def _file_state(path: Path) -> tuple[bytes, int]:
    stat = path.stat()
    return path.read_bytes(), stat.st_mtime_ns


def _tree_state() -> tuple[tuple[object, ...], ...]:
    state = []
    for root in GENERATED_ROOTS:
        state.append((root, root.exists()))
        if root.exists():
            for path in sorted(root.rglob("*")):
                state.append(
                    (
                        path,
                        path.is_dir(),
                        path.is_symlink(),
                        path.read_bytes() if path.is_file() else None,
                    )
                )
    return tuple(state)


def _copy_synthetic_pair(root: Path) -> Path:
    incoming = root / "incoming"
    incoming.mkdir(parents=True)
    shutil.copyfile(
        SYNTHETIC_GLOSSARY_WORKBOOK,
        incoming / GLOSSARY_WORKBOOK_FILENAME,
    )
    shutil.copyfile(
        SYNTHETIC_REFERENCES_WORKBOOK,
        incoming / REFERENCES_WORKBOOK_FILENAME,
    )
    return incoming


def _edit_workbook(path: Path, edit) -> None:
    workbook = load_workbook(path)
    try:
        edit(workbook)
        workbook.save(path)
    finally:
        workbook.close()


def _contract(incoming: Path, root: Path):
    publication = publish_or_reuse_intake(
        inventory_incoming_pair(read_incoming_pair(incoming)),
        root / "raw",
        root / "manifest",
        source_git=GitSourceState("unknown", None),
    )
    return validate_workbook_contract(publication)


def _synthetic_registry() -> tuple[RegistryEntry, ...]:
    return (
        RegistryEntry(
            "RHC-901",
            901,
            "Synthetic Alpha",
            RegistryStatus.ACTIVE,
            None,
            "20000115",
            None,
        ),
        RegistryEntry(
            "RHC-902",
            902,
            "Synthetic Beta",
            RegistryStatus.ACTIVE,
            None,
            "20000115",
            None,
        ),
        RegistryEntry(
            "RHC-903",
            903,
            "Synthetic Mixture",
            RegistryStatus.ACTIVE,
            None,
            "20000115",
            None,
        ),
    )


def _synthetic_crosswalk() -> tuple[TrackedCrosswalkEntry, ...]:
    return (
        TrackedCrosswalkEntry(
            "Synthetic Alpha",
            "RHC-901",
            ReferenceResolutionMethod.EXACT,
            "20000115",
        ),
        TrackedCrosswalkEntry(
            "Synthetic Beta",
            "RHC-902",
            ReferenceResolutionMethod.OVERRIDE,
            "20000115",
        ),
        TrackedCrosswalkEntry(
            "Synthetic Mixture",
            "RHC-903",
            ReferenceResolutionMethod.EXACT,
            "20000115",
        ),
    )


def _write_assets(
    root: Path,
    registry: tuple[RegistryEntry, ...] | None = None,
    crosswalk: tuple[TrackedCrosswalkEntry, ...] | None = None,
) -> tuple[Path, Path]:
    registry_values = registry if registry is not None else _synthetic_registry()
    crosswalk_values = (
        crosswalk if crosswalk is not None else _synthetic_crosswalk()
    )
    registry_path = root / "contaminant_registry.csv"
    crosswalk_path = root / "reference_crosswalk.csv"
    registry_path.write_bytes(serialize_registry(registry_values))
    crosswalk_path.write_bytes(
        serialize_crosswalk(crosswalk_values, registry_values)
    )
    return registry_path, crosswalk_path


def _finding_codes(error: IdentityRelationshipValidationError) -> set[str]:
    return {finding.code for finding in error.findings}


class IdentityRelationshipValidationTests(unittest.TestCase):
    def test_release_order_is_shared_and_numeric_for_same_day_revisions(self) -> None:
        self.assertLess(release_order_key("20000115"), release_order_key("20000115-r2"))
        self.assertLess(
            release_order_key("20000115-r2"), release_order_key("20000115-r10")
        )
        self.assertLess(
            release_order_key("20000115-r10"), release_order_key("20000116")
        )

    def test_resolves_ids_references_and_footnotes_from_raw_snapshots(self) -> None:
        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            incoming = _copy_synthetic_pair(root)
            contract = _contract(incoming, root)
            registry_path, crosswalk_path = _write_assets(root)
            protected = {
                path: _file_state(path)
                for path in (
                    contract.raw_pair.glossary_snapshot.path,
                    contract.raw_pair.references_snapshot.path,
                    registry_path,
                    crosswalk_path,
                )
            }
            shutil.rmtree(incoming)

            result = validate_identity_relationships(
                contract, registry_path, crosswalk_path
            )

            self.assertEqual(result.data_release_id, "20000115-r2")
            self.assertEqual(result.schema_version, "1.0.0")
            self.assertEqual(
                [value.id_contaminant for value in result.glossary_identities],
                ["RHC-901", "RHC-902", "RHC-903"],
            )
            self.assertEqual(len(result.reference_relationships), 4)
            self.assertEqual(
                result.reference_relationships[2].resolution_method,
                ReferenceResolutionMethod.OVERRIDE,
            )
            self.assertEqual(len(result.footnote_definitions), 2)
            self.assertEqual(len(result.footnote_usages), 3)
            self.assertEqual(result.duplicate_candidates, ())
            self.assertEqual(result.findings, ())
            self.assertEqual(
                {path: _file_state(path) for path in protected}, protected
            )
            with self.assertRaises(FrozenInstanceError):
                result.data_release_id = "changed"

    def test_future_registry_and_crosswalk_rows_are_ignored_historically(self) -> None:
        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            contract = _contract(_copy_synthetic_pair(root), root)
            registry = _synthetic_registry() + (
                RegistryEntry(
                    "RHC-904",
                    None,
                    "Future identity",
                    RegistryStatus.ACTIVE,
                    None,
                    "20000116",
                    None,
                ),
            )
            crosswalk = _synthetic_crosswalk() + (
                TrackedCrosswalkEntry(
                    "Future label",
                    "RHC-904",
                    ReferenceResolutionMethod.OVERRIDE,
                    "20000116",
                ),
            )
            registry_path, crosswalk_path = _write_assets(
                root, registry, crosswalk
            )

            result = validate_identity_relationships(
                contract, registry_path, crosswalk_path
            )

            self.assertEqual(len(result.registry_entries), 3)
            self.assertEqual(len(result.crosswalk_entries), 3)

    def test_active_schema_one_identity_requires_a_legacy_id(self) -> None:
        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            contract = _contract(_copy_synthetic_pair(root), root)
            registry = _synthetic_registry() + (
                RegistryEntry(
                    "RHC-904",
                    None,
                    "Unaddressable current identity",
                    RegistryStatus.ACTIVE,
                    None,
                    "20000115",
                    None,
                ),
            )
            registry_path, crosswalk_path = _write_assets(root, registry)

            with self.assertRaises(IdentityRelationshipValidationError) as caught:
                validate_identity_relationships(
                    contract, registry_path, crosswalk_path
                )

            codes = _finding_codes(caught.exception)
            self.assertIn("active_identity_without_legacy_id", codes)
            self.assertIn("missing_active_identity", codes)

    def test_rejects_wrong_input_and_malformed_assets(self) -> None:
        with self.assertRaises(IdentityRelationshipValidationError) as caught:
            validate_identity_relationships(object())
        self.assertEqual(_finding_codes(caught.exception), {"invalid_workbook_contract"})

        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            contract = _contract(_copy_synthetic_pair(root), root)
            registry_path, crosswalk_path = _write_assets(root)
            registry_path.write_bytes(b"not,a,registry\n")

            with self.assertRaises(IdentityRelationshipValidationError) as caught:
                validate_identity_relationships(
                    contract, registry_path, crosswalk_path
                )

            self.assertIn("invalid_registry_asset", _finding_codes(caught.exception))

        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            contract = _contract(_copy_synthetic_pair(root), root)
            registry_path, crosswalk_path = _write_assets(root)
            crosswalk_path.write_bytes(b"not,a,crosswalk\n")

            with self.assertRaises(IdentityRelationshipValidationError) as caught:
                validate_identity_relationships(
                    contract, registry_path, crosswalk_path
                )

            self.assertIn("invalid_crosswalk_asset", _finding_codes(caught.exception))

    def test_rejects_invalid_duplicate_missing_and_retired_identities(self) -> None:
        cases = (
            ("blank", None, "invalid_legacy_id"),
            ("boolean", True, "invalid_legacy_id"),
            ("text", "901", "invalid_legacy_id"),
            ("float", 901.5, "invalid_legacy_id"),
            ("zero", 0, "invalid_legacy_id"),
            ("negative", -1, "invalid_legacy_id"),
            ("duplicate", 902, "duplicate_legacy_id"),
            ("absent", 999, "legacy_id_not_in_applicable_registry"),
        )
        for name, replacement, code in cases:
            with self.subTest(name=name), TemporaryDirectory() as temporary:
                root = Path(temporary)
                incoming = _copy_synthetic_pair(root)
                _edit_workbook(
                    incoming / GLOSSARY_WORKBOOK_FILENAME,
                    lambda workbook, value=replacement: setattr(
                        workbook[GLOSSARY_SHEET_NAME]["C2"], "value", value
                    ),
                )
                contract = _contract(incoming, root)
                registry_path, crosswalk_path = _write_assets(root)

                with self.assertRaises(IdentityRelationshipValidationError) as caught:
                    validate_identity_relationships(
                        contract, registry_path, crosswalk_path
                    )

                self.assertIn(code, _finding_codes(caught.exception))
                self.assertIn("missing_active_identity", _finding_codes(caught.exception))

        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            contract = _contract(_copy_synthetic_pair(root), root)
            registry = (
                RegistryEntry(
                    "RHC-901",
                    901,
                    "Synthetic Alpha",
                    RegistryStatus.RETIRED,
                    None,
                    "20000115",
                    "20000115-r2",
                ),
            ) + _synthetic_registry()[1:]
            registry_path, crosswalk_path = _write_assets(root, registry)

            with self.assertRaises(IdentityRelationshipValidationError) as caught:
                validate_identity_relationships(
                    contract, registry_path, crosswalk_path
                )

            self.assertIn("retired_identity_present", _finding_codes(caught.exception))
            self.assertIn("reference_target_not_active", _finding_codes(caught.exception))

    def test_rejects_formula_identity_and_reference_cells(self) -> None:
        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            incoming = _copy_synthetic_pair(root)
            _edit_workbook(
                incoming / GLOSSARY_WORKBOOK_FILENAME,
                lambda workbook: setattr(
                    workbook[GLOSSARY_SHEET_NAME]["C2"], "value", "=900+1"
                ),
            )
            _edit_workbook(
                incoming / REFERENCES_WORKBOOK_FILENAME,
                lambda workbook: setattr(
                    workbook[REFERENCES_SHEET_NAME]["A2"],
                    "value",
                    '=CONCAT("Synthetic ","Alpha")',
                ),
            )
            contract = _contract(incoming, root)
            registry_path, crosswalk_path = _write_assets(root)

            with self.assertRaises(IdentityRelationshipValidationError) as caught:
                validate_identity_relationships(
                    contract, registry_path, crosswalk_path
                )

            codes = _finding_codes(caught.exception)
            self.assertIn("formula_legacy_id", codes)
            self.assertIn("formula_reference_label", codes)

    def test_reference_labels_are_exact_and_repeated_rows_are_allowed(self) -> None:
        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            incoming = _copy_synthetic_pair(root)
            contract = _contract(incoming, root)
            registry_path, crosswalk_path = _write_assets(root)
            result = validate_identity_relationships(
                contract, registry_path, crosswalk_path
            )
            alpha_rows = [
                row
                for row in result.reference_relationships
                if row.refs_review_name == "Synthetic Alpha"
            ]
            self.assertEqual(len(alpha_rows), 2)

            incoming_variant = _copy_synthetic_pair(root / "variant")
            _edit_workbook(
                incoming_variant / REFERENCES_WORKBOOK_FILENAME,
                lambda workbook: setattr(
                    workbook[REFERENCES_SHEET_NAME]["A2"],
                    "value",
                    "synthetic alpha",
                ),
            )
            variant_contract = _contract(incoming_variant, root / "variant")
            with self.assertRaises(IdentityRelationshipValidationError) as caught:
                validate_identity_relationships(
                    variant_contract, registry_path, crosswalk_path
                )
            self.assertIn("unresolved_reference_label", _finding_codes(caught.exception))

    def test_reference_validation_rejects_blank_nontext_and_future_only_labels(self) -> None:
        cases = (("blank", None), ("nontext", 17))
        for name, replacement in cases:
            with self.subTest(name=name), TemporaryDirectory() as temporary:
                root = Path(temporary)
                incoming = _copy_synthetic_pair(root)
                _edit_workbook(
                    incoming / REFERENCES_WORKBOOK_FILENAME,
                    lambda workbook, value=replacement: setattr(
                        workbook[REFERENCES_SHEET_NAME]["A2"], "value", value
                    ),
                )
                contract = _contract(incoming, root)
                registry_path, crosswalk_path = _write_assets(root)
                with self.assertRaises(IdentityRelationshipValidationError) as caught:
                    validate_identity_relationships(
                        contract, registry_path, crosswalk_path
                    )
                self.assertIn("invalid_reference_label", _finding_codes(caught.exception))

        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            contract = _contract(_copy_synthetic_pair(root), root)
            future_crosswalk = tuple(
                TrackedCrosswalkEntry(
                    entry.refs_review_name,
                    entry.id_contaminant,
                    entry.resolution_method,
                    "20000116"
                    if entry.refs_review_name == "Synthetic Alpha"
                    else entry.reviewed_release_id,
                )
                for entry in _synthetic_crosswalk()
            )
            registry_path, crosswalk_path = _write_assets(
                root, crosswalk=future_crosswalk
            )
            with self.assertRaises(IdentityRelationshipValidationError) as caught:
                validate_identity_relationships(
                    contract, registry_path, crosswalk_path
                )
            self.assertIn("unresolved_reference_label", _finding_codes(caught.exception))

    def test_name_differences_are_review_findings_not_identity_changes(self) -> None:
        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            incoming = _copy_synthetic_pair(root)
            _edit_workbook(
                incoming / REFERENCES_WORKBOOK_FILENAME,
                lambda workbook: setattr(
                    workbook[REFERENCES_SHEET_NAME]["A2"],
                    "value",
                    "Reviewed Alpha label",
                ),
            )
            contract = _contract(incoming, root)
            registry = (
                RegistryEntry(
                    "RHC-901",
                    901,
                    "Registry Alpha name",
                    RegistryStatus.ACTIVE,
                    None,
                    "20000115",
                    None,
                ),
            ) + _synthetic_registry()[1:]
            crosswalk = (
                TrackedCrosswalkEntry(
                    "Reviewed Alpha label",
                    "RHC-901",
                    ReferenceResolutionMethod.OVERRIDE,
                    "20000115",
                ),
            ) + _synthetic_crosswalk()
            registry_path, crosswalk_path = _write_assets(
                root, registry, crosswalk
            )

            result = validate_identity_relationships(
                contract, registry_path, crosswalk_path
            )

            self.assertEqual(result.glossary_identities[0].id_contaminant, "RHC-901")
            self.assertEqual(
                {finding.code for finding in result.findings},
                {
                    "registry_name_mismatch",
                    "reference_label_differs_from_glossary_name",
                },
            )

    def test_identity_resolution_is_independent_of_glossary_row_order(self) -> None:
        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            incoming = _copy_synthetic_pair(root)

            def swap_rows(workbook) -> None:
                sheet = workbook[GLOSSARY_SHEET_NAME]
                second = [cell.value for cell in sheet[2]]
                third = [cell.value for cell in sheet[3]]
                for column, value in enumerate(third, start=1):
                    sheet.cell(2, column).value = value
                for column, value in enumerate(second, start=1):
                    sheet.cell(3, column).value = value

            _edit_workbook(incoming / GLOSSARY_WORKBOOK_FILENAME, swap_rows)
            contract = _contract(incoming, root)
            registry_path, crosswalk_path = _write_assets(root)

            result = validate_identity_relationships(
                contract, registry_path, crosswalk_path
            )

            self.assertEqual(
                [identity.id_contaminant for identity in result.glossary_identities],
                ["RHC-901", "RHC-902", "RHC-903"],
            )

    def test_footnote_problems_join_into_the_shared_error_model(self) -> None:
        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            incoming = _copy_synthetic_pair(root)
            _edit_workbook(
                incoming / GLOSSARY_WORKBOOK_FILENAME,
                lambda workbook: setattr(
                    workbook[GLOSSARY_SHEET_NAME]["AR2"], "value", "UNKNOWN"
                ),
            )
            contract = _contract(incoming, root)
            registry_path, crosswalk_path = _write_assets(root)

            with self.assertRaises(IdentityRelationshipValidationError) as caught:
                validate_identity_relationships(
                    contract, registry_path, crosswalk_path
                )

            footnote_findings = [
                finding
                for finding in caught.exception.findings
                if finding.code == "invalid_footnote_relationship"
            ]
            self.assertEqual(len(footnote_findings), 1)
            self.assertEqual(footnote_findings[0].sheet, GLOSSARY_SHEET_NAME)

    def test_reports_exact_duplicate_candidates_without_merging(self) -> None:
        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            incoming = _copy_synthetic_pair(root)

            def add_duplicates(workbook) -> None:
                sheet = workbook[GLOSSARY_SHEET_NAME]
                sheet["A3"] = "Synthetic Alpha"
                sheet["F3"] = "100-00-1 | 100-00-1"
                sheet["G3"] = "AAAAAAAAAAAAAA-BBBBBBBBBB-C"
                sheet["F4"] = "NA"
                sheet["G4"] = "N/A"

            _edit_workbook(
                incoming / GLOSSARY_WORKBOOK_FILENAME, add_duplicates
            )
            contract = _contract(incoming, root)
            registry_path, crosswalk_path = _write_assets(root)

            result = validate_identity_relationships(
                contract, registry_path, crosswalk_path
            )

            self.assertEqual(len(result.glossary_identities), 3)
            self.assertEqual(
                {candidate.canonical_field for candidate in result.duplicate_candidates},
                {"id_name", "id_casrn", "id_inchikey"},
            )
            self.assertTrue(
                all(
                    finding.severity is ValidationSeverity.WARNING
                    for finding in result.findings
                    if finding.category.value == "duplicates"
                )
            )
            self.assertNotIn(
                "NA", {value.source_value for value in result.duplicate_candidates}
            )
            self.assertNotIn(
                "N/A", {value.source_value for value in result.duplicate_candidates}
            )


class AuthoritativeIdentityRelationshipValidationTests(unittest.TestCase):
    def test_validates_authoritative_relationships_without_modifying_sources(self) -> None:
        protected_paths = (
            INCOMING_GLOSSARY_WORKBOOK,
            INCOMING_REFERENCES_WORKBOOK,
            CONTAMINANT_REGISTRY_PATH,
            REFERENCE_CROSSWALK_PATH,
            PIPELINE_ROOT / "pyproject.toml",
            PIPELINE_ROOT / "uv.lock",
        )
        protected_before = {path: _file_state(path) for path in protected_paths}
        generated_before = _tree_state()

        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            contract = _contract(INCOMING_DIR, root)
            result = validate_identity_relationships(contract)

            self.assertEqual(len(result.glossary_identities), 152)
            self.assertEqual(len(result.reference_relationships), 406)
            self.assertEqual(
                len({row.refs_review_name for row in result.reference_relationships}),
                133,
            )
            self.assertEqual(len(result.footnote_definitions), 4)
            self.assertEqual(
                sum(bool(usage.footnote_ids) for usage in result.footnote_usages),
                33,
            )
            info = [
                finding
                for finding in result.findings
                if finding.severity is ValidationSeverity.INFO
            ]
            warnings = [
                finding
                for finding in result.findings
                if finding.severity is ValidationSeverity.WARNING
            ]
            self.assertEqual(len(info), 21)
            self.assertEqual(len(warnings), 3)
            self.assertEqual(len(result.duplicate_candidates), 3)
            self.assertEqual(
                {candidate.id_contaminants for candidate in result.duplicate_candidates},
                {
                    ("RHC-020", "RHC-037"),
                    ("RHC-106", "RHC-107"),
                },
            )

        self.assertEqual(
            {path: _file_state(path) for path in protected_paths}, protected_before
        )
        self.assertEqual(_tree_state(), generated_before)


if __name__ == "__main__":
    unittest.main()
