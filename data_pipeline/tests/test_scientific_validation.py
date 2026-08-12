"""Tests for literal scientific and source-value validation."""

from dataclasses import FrozenInstanceError, replace
from pathlib import Path
import shutil
from tempfile import TemporaryDirectory
import unittest

from openpyxl import load_workbook

from contaminant_pipeline.config import (
    GLOSSARY_SHEET_NAME,
    GLOSSARY_WORKBOOK_FILENAME,
    REFERENCES_SHEET_NAME,
    REFERENCES_WORKBOOK_FILENAME,
)
from contaminant_pipeline.crosswalk import ReferenceResolutionMethod
from contaminant_pipeline.intake import (
    GitSourceState,
    inventory_incoming_pair,
    publish_or_reuse_intake,
    read_incoming_pair,
)
from contaminant_pipeline.paths import (
    CONTAMINANT_REGISTRY_PATH,
    INCOMING_DIR,
    INCOMING_GLOSSARY_WORKBOOK,
    INCOMING_REFERENCES_WORKBOOK,
    PIPELINE_ROOT,
    REFERENCE_CROSSWALK_PATH,
)
from contaminant_pipeline.registry_assets import (
    RegistryEntry,
    RegistryStatus,
    TrackedCrosswalkEntry,
    serialize_crosswalk,
    serialize_registry,
)
from contaminant_pipeline.schemas import NotApplicable
from contaminant_pipeline.scientific_validation import (
    ScientificFieldValidationError,
    validate_scientific_fields,
)
from contaminant_pipeline.validate import (
    validate_identity_relationships,
    validate_workbook_contract,
)

from fixture_paths import (
    SYNTHETIC_GLOSSARY_WORKBOOK,
    SYNTHETIC_REFERENCES_WORKBOOK,
)


GENERATED_ROOTS = (
    PIPELINE_ROOT / "data" / "01_manifest",
    PIPELINE_ROOT / "data" / "02_raw_snapshots",
    PIPELINE_ROOT / "data" / "03_processed",
    PIPELINE_ROOT / "data" / "04_output",
    PIPELINE_ROOT.parent / "public" / "data",
)


def _file_state(path: Path) -> tuple[bytes, int]:
    stat = path.stat()
    return path.read_bytes(), stat.st_mtime_ns


def _tree_state() -> tuple[tuple[object, ...], ...]:
    state = []
    for root in GENERATED_ROOTS:
        state.append((root, root.exists()))
        if root.exists():
            for path in sorted(root.rglob("*")):
                state.append(
                    (
                        path,
                        path.is_dir(),
                        path.is_symlink(),
                        path.read_bytes() if path.is_file() else None,
                    )
                )
    return tuple(state)


def _copy_synthetic_pair(root: Path) -> Path:
    incoming = root / "incoming"
    incoming.mkdir(parents=True)
    shutil.copyfile(
        SYNTHETIC_GLOSSARY_WORKBOOK,
        incoming / GLOSSARY_WORKBOOK_FILENAME,
    )
    shutil.copyfile(
        SYNTHETIC_REFERENCES_WORKBOOK,
        incoming / REFERENCES_WORKBOOK_FILENAME,
    )
    return incoming


def _edit_workbook(path: Path, edit) -> None:
    workbook = load_workbook(path)
    try:
        edit(workbook)
        workbook.save(path)
    finally:
        workbook.close()


def _set_glossary_values(path: Path, changes: tuple[tuple[int, str, object], ...]):
    def edit(workbook) -> None:
        sheet = workbook[GLOSSARY_SHEET_NAME]
        columns = {cell.value: cell.column for cell in sheet[1]}
        for source_row, header, value in changes:
            sheet.cell(source_row, columns[header]).value = value

    _edit_workbook(path, edit)


def _set_reference_values(path: Path, changes: tuple[tuple[int, str, object], ...]):
    def edit(workbook) -> None:
        sheet = workbook[REFERENCES_SHEET_NAME]
        columns = {cell.value: cell.column for cell in sheet[1]}
        for source_row, header, value in changes:
            sheet.cell(source_row, columns[header]).value = value

    _edit_workbook(path, edit)


def _make_scientifically_valid(incoming: Path) -> None:
    _set_glossary_values(
        incoming / GLOSSARY_WORKBOOK_FILENAME,
        (
            (2, "Chemical formula", "C2H6O"),
            (2, "CASRN", "64-17-5 | 67-56-1"),
            (2, "F", 0),
            (3, "Chemical formula", "C6H6"),
        ),
    )


def _contract(incoming: Path, root: Path):
    publication = publish_or_reuse_intake(
        inventory_incoming_pair(read_incoming_pair(incoming)),
        root / "raw",
        root / "manifest",
        source_git=GitSourceState("unknown", None),
    )
    return validate_workbook_contract(publication)


def _synthetic_registry(alpha_id: str = "RHC-901") -> tuple[RegistryEntry, ...]:
    return (
        RegistryEntry(
            alpha_id,
            901,
            "Synthetic Alpha",
            RegistryStatus.ACTIVE,
            None,
            "20000115",
            None,
        ),
        RegistryEntry(
            "RHC-902",
            902,
            "Synthetic Beta",
            RegistryStatus.ACTIVE,
            None,
            "20000115",
            None,
        ),
        RegistryEntry(
            "RHC-903",
            903,
            "Synthetic Mixture",
            RegistryStatus.ACTIVE,
            None,
            "20000115",
            None,
        ),
    )


def _synthetic_crosswalk(alpha_id: str = "RHC-901"):
    return (
        TrackedCrosswalkEntry(
            "Synthetic Alpha",
            alpha_id,
            ReferenceResolutionMethod.EXACT,
            "20000115",
        ),
        TrackedCrosswalkEntry(
            "Synthetic Beta",
            "RHC-902",
            ReferenceResolutionMethod.OVERRIDE,
            "20000115",
        ),
        TrackedCrosswalkEntry(
            "Synthetic Mixture",
            "RHC-903",
            ReferenceResolutionMethod.EXACT,
            "20000115",
        ),
    )


def _write_assets(root: Path, alpha_id: str = "RHC-901") -> tuple[Path, Path]:
    registry = _synthetic_registry(alpha_id)
    crosswalk = _synthetic_crosswalk(alpha_id)
    registry_path = root / "contaminant_registry.csv"
    crosswalk_path = root / "reference_crosswalk.csv"
    registry_path.write_bytes(serialize_registry(registry))
    crosswalk_path.write_bytes(serialize_crosswalk(crosswalk, registry))
    return registry_path, crosswalk_path


def _identity_result(incoming: Path, root: Path, alpha_id: str = "RHC-901"):
    registry_path, crosswalk_path = _write_assets(root, alpha_id)
    return validate_identity_relationships(
        _contract(incoming, root), registry_path, crosswalk_path
    )


def _finding_codes(error: ScientificFieldValidationError) -> set[str]:
    return {finding.code for finding in error.findings}


class ScientificFieldValidationTests(unittest.TestCase):
    def test_validates_and_freezes_scientific_values_from_raw_snapshots(self) -> None:
        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            incoming = _copy_synthetic_pair(root)
            _make_scientifically_valid(incoming)
            identities = _identity_result(incoming, root)
            raw_paths = (
                identities.workbook_contract.raw_pair.glossary_snapshot.path,
                identities.workbook_contract.raw_pair.references_snapshot.path,
            )
            protected = {path: _file_state(path) for path in raw_paths}
            shutil.rmtree(incoming)

            result = validate_scientific_fields(identities)

            self.assertEqual(result.data_release_id, "20000115-r2")
            self.assertEqual(len(result.contaminants), 3)
            self.assertEqual(len(result.references), 4)
            by_id = {row.id_contaminant: row for row in result.contaminants}
            self.assertEqual(
                by_id["RHC-901"].parsed_values["id_casrn"],
                ("64-17-5", "67-56-1"),
            )
            self.assertEqual(
                by_id["RHC-903"].parsed_values["id_chem_formula"],
                NotApplicable.VALUE,
            )
            with self.assertRaises(TypeError):
                by_id["RHC-901"].parsed_values["id_casrn"] = ()
            with self.assertRaises(FrozenInstanceError):
                result.data_release_id = "changed"
            self.assertEqual(
                {path: _file_state(path) for path in raw_paths}, protected
            )

    def test_rejects_wrong_stage_input(self) -> None:
        with self.assertRaises(ScientificFieldValidationError) as raised:
            validate_scientific_fields(object())
        self.assertEqual(
            _finding_codes(raised.exception), {"invalid_identity_relationships"}
        )

    def test_validates_casrn_inchikey_formula_and_literal_delimiters(self) -> None:
        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            incoming = _copy_synthetic_pair(root)
            _make_scientifically_valid(incoming)
            _set_glossary_values(
                incoming / GLOSSARY_WORKBOOK_FILENAME,
                (
                    (2, "CASRN", "64-17-4"),
                    (2, "InChIKey", "lowercase-invalid-key"),
                    (2, "Chemical formula", "X2"),
                    (2, "a.k.a.s", "Alpha|Alias"),
                ),
            )
            identities = _identity_result(incoming, root)

            with self.assertRaises(ScientificFieldValidationError) as raised:
                validate_scientific_fields(identities)

            self.assertTrue(
                {
                    "invalid_id_casrn",
                    "invalid_id_inchikey",
                    "invalid_id_chem_formula",
                    "invalid_id_aka",
                }.issubset(_finding_codes(raised.exception))
            )

    def test_not_applicable_requires_classification_or_stable_exception(self) -> None:
        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            incoming = _copy_synthetic_pair(root)
            _make_scientifically_valid(incoming)
            _set_glossary_values(
                incoming / GLOSSARY_WORKBOOK_FILENAME,
                ((2, "CASRN", "NA"),),
            )
            identities = _identity_result(incoming, root)
            with self.assertRaises(ScientificFieldValidationError) as raised:
                validate_scientific_fields(identities)
            self.assertIn("invalid_id_casrn", _finding_codes(raised.exception))

        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            incoming = _copy_synthetic_pair(root)
            _make_scientifically_valid(incoming)
            _set_glossary_values(
                incoming / GLOSSARY_WORKBOOK_FILENAME,
                ((2, "InChIKey", "NA"),),
            )
            identities = _identity_result(incoming, root, alpha_id="RHC-071")
            result = validate_scientific_fields(identities)
            by_id = {row.id_contaminant: row for row in result.contaminants}
            self.assertEqual(
                by_id["RHC-071"].parsed_values["id_inchikey"],
                NotApplicable.VALUE,
            )

    def test_mixtures_may_use_not_applicable_or_real_identifiers(self) -> None:
        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            incoming = _copy_synthetic_pair(root)
            _make_scientifically_valid(incoming)
            _set_glossary_values(
                incoming / GLOSSARY_WORKBOOK_FILENAME,
                (
                    (4, "CASRN", "NA"),
                    (4, "InChIKey", "N/A"),
                ),
            )
            identities = _identity_result(incoming, root)
            result = validate_scientific_fields(identities)
            mixture = next(
                row for row in result.contaminants if row.id_contaminant == "RHC-903"
            )
            self.assertEqual(mixture.parsed_values["id_casrn"], NotApplicable.VALUE)
            self.assertEqual(
                mixture.parsed_values["id_inchikey"], NotApplicable.VALUE
            )

    def test_rejects_invalid_types_ranges_regulatory_values_and_enums(self) -> None:
        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            incoming = _copy_synthetic_pair(root)
            _make_scientifically_valid(incoming)
            _set_glossary_values(
                incoming / GLOSSARY_WORKBOOK_FILENAME,
                (
                    (2, "Primary", "Organic"),
                    (2, "Aromatic", 1),
                    (2, "C", "3 - 1"),
                    (2, "Ethers", -1),
                    (2, "Halogenated", "NA"),
                    (2, "NPDWR MCL (mg/L)", -0.1),
                    (2, "SMCL (mg/L)", "7-9"),
                    (2, "HDOH MCL (mg/L)", "AL = 2"),
                    (2, "Stockholm Convention", "B"),
                ),
            )
            identities = _identity_result(incoming, root)

            with self.assertRaises(ScientificFieldValidationError) as raised:
                validate_scientific_fields(identities)

            self.assertTrue(
                {
                    "invalid_class_primary",
                    "invalid_class_aromatic",
                    "invalid_chem_info_n_carbon",
                    "invalid_chem_info_n_ether",
                    "invalid_chem_info_halogenated",
                    "invalid_reg_status_npdwr_mcl_mg_l",
                    "invalid_reg_status_smcl",
                    "invalid_reg_status_hdoh_mcl_mg_l",
                    "invalid_reg_status_stockholm_convention",
                }.issubset(_finding_codes(raised.exception))
            )

    def test_rejects_text_and_url_problems_and_aggregates_in_order(self) -> None:
        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            incoming = _copy_synthetic_pair(root)
            _make_scientifically_valid(incoming)
            _set_glossary_values(
                incoming / GLOSSARY_WORKBOOK_FILENAME,
                ((2, "Sources", " trailing "),),
            )
            _set_reference_values(
                incoming / REFERENCES_WORKBOOK_FILENAME,
                (
                    (2, "source", ""),
                    (3, "link", "ftp://example.com/not-http"),
                ),
            )
            identities = _identity_result(incoming, root)

            with self.assertRaises(ScientificFieldValidationError) as raised:
                validate_scientific_fields(identities)

            self.assertEqual(
                [
                    finding.code
                    for finding in raised.exception.findings
                    if finding.severity.value == "error"
                ],
                [
                    "invalid_refs_source",
                    "invalid_refs_url",
                    "invalid_source_notes_sources",
                ],
            )

    def test_retains_formula_context_for_3_4_without_rejecting_cached_value(self) -> None:
        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            incoming = _copy_synthetic_pair(root)
            _make_scientifically_valid(incoming)
            identities = _identity_result(incoming, root)
            contract = identities.workbook_contract
            snapshot = contract.raw_pair.glossary_snapshot
            revised_sheets = []
            for sheet in snapshot.sheets:
                if sheet.name != GLOSSARY_SHEET_NAME:
                    revised_sheets.append(sheet)
                    continue
                revised_sheets.append(
                    replace(
                        sheet,
                        cells=tuple(
                            replace(cell, formula="=\"C2H6O\"")
                            if cell.coordinate == "D2"
                            else cell
                            for cell in sheet.cells
                        ),
                    )
                )
            revised_snapshot = replace(snapshot, sheets=tuple(revised_sheets))
            revised_pair = replace(
                contract.raw_pair, glossary_snapshot=revised_snapshot
            )
            revised_contract = replace(contract, raw_pair=revised_pair)
            revised_identities = replace(
                identities, workbook_contract=revised_contract
            )

            result = validate_scientific_fields(revised_identities)

            alpha = next(
                row for row in result.contaminants if row.id_contaminant == "RHC-901"
            )
            self.assertEqual(alpha.formulas["id_chem_formula"], '=\"C2H6O\"')
            self.assertEqual(alpha.parsed_values["id_chem_formula"], "C2H6O")


class AuthoritativeScientificFieldValidationTests(unittest.TestCase):
    def test_reports_authoritative_defects_without_modifying_sources(self) -> None:
        protected_paths = (
            INCOMING_GLOSSARY_WORKBOOK,
            INCOMING_REFERENCES_WORKBOOK,
            CONTAMINANT_REGISTRY_PATH,
            REFERENCE_CROSSWALK_PATH,
            PIPELINE_ROOT / "pyproject.toml",
            PIPELINE_ROOT / "uv.lock",
        )
        protected_before = {path: _file_state(path) for path in protected_paths}
        generated_before = _tree_state()

        with TemporaryDirectory() as temporary:
            root = Path(temporary)
            identities = validate_identity_relationships(_contract(INCOMING_DIR, root))
            with self.assertRaises(ScientificFieldValidationError) as raised:
                validate_scientific_fields(identities)

            findings = raised.exception.findings
            self.assertEqual(len(findings), 63)
            counts = {}
            for finding in findings:
                key = (finding.severity.value, finding.code)
                counts[key] = counts.get(key, 0) + 1
            self.assertEqual(
                counts,
                {
                    ("error", "invalid_id_casrn"): 1,
                    ("error", "invalid_id_chem_formula"): 5,
                    ("warning", "pending_id_casrn"): 16,
                    ("warning", "pending_id_inchikey"): 16,
                    ("warning", "pending_source_notes_sources"): 6,
                    ("warning", "unverified_id_casrn_not_applicable"): 9,
                    ("warning", "unverified_id_inchikey_not_applicable"): 10,
                },
            )
            self.assertIn("invalid_id_casrn", _finding_codes(raised.exception))
            casrn = next(
                finding
                for finding in findings
                if finding.code == "invalid_id_casrn"
                and finding.id_contaminant == "RHC-015"
            )
            self.assertEqual(casrn.source_value, "'207-916-6'")

        self.assertEqual(
            {path: _file_state(path) for path in protected_paths}, protected_before
        )
        self.assertEqual(_tree_state(), generated_before)


if __name__ == "__main__":
    unittest.main()
