"""Read, inventory, snapshot, manifest, and safely reconcile workbook intake."""

from collections import Counter
from dataclasses import dataclass
from hashlib import sha256
import json
import os
from pathlib import Path
import re
import shutil
import subprocess
import tempfile

from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries

from .config import (
    FOOTNOTES_HEADER_ROW,
    FOOTNOTES_SHEET_NAME,
    GLOSSARY_SHEET_NAME,
    GLOSSARY_TABLE_NAME,
    GLOSSARY_WORKBOOK_FILENAME,
    GLOSSARY_WORKBOOK_TYPE,
    GLOSSARY_WORKSHEET_NAMES,
    INTAKE_MANIFEST_SCHEMA_VERSION,
    INTRODUCTION_SHEET_NAME,
    METADATA_SHEET_NAME,
    METADATA_TABLE_NAME,
    REFERENCES_HEADER_ROW,
    REFERENCES_SHEET_NAME,
    REFERENCES_WORKBOOK_FILENAME,
    REFERENCES_WORKBOOK_TYPE,
    REFERENCES_WORKSHEET_NAMES,
    validate_release_id,
)
from .io_excel import (
    CellSnapshot,
    ExcelReadError,
    TableSnapshot,
    WorkbookSnapshot,
    WorksheetSnapshot,
    read_workbook,
)
from .metadata import (
    WorkbookCompatibility,
    WorkbookMetadata,
    extract_workbook_metadata,
    validate_workbook_compatibility,
)
from .paths import (
    INCOMING_DIR,
    MANIFEST_DIR,
    RAW_SNAPSHOTS_DIR,
    REPOSITORY_ROOT,
)


@dataclass(frozen=True)
class IncomingWorkbookPair:
    """Two read-only workbook snapshots with validated release identity."""

    glossary_snapshot: WorkbookSnapshot
    references_snapshot: WorkbookSnapshot
    compatibility: WorkbookCompatibility


class IncomingContractError(ValueError):
    """Raised when the stable incoming workbook contract is not satisfied."""


@dataclass(frozen=True)
class HeaderCellInventory:
    """One literal header and its source position."""

    column: int
    coordinate: str
    value: str


@dataclass(frozen=True)
class FormulaInventory:
    """One formula definition and cached-value state."""

    worksheet: str
    coordinate: str
    definition: str
    has_cached_value: bool


@dataclass(frozen=True)
class TableInventory:
    """Deterministic structure and row counts for one Excel table."""

    name: str
    reference: str
    min_row: int
    max_row: int
    min_column: int
    max_column: int
    header_row: int
    headers: tuple[HeaderCellInventory, ...]
    declared_data_row_count: int
    populated_data_row_count: int


@dataclass(frozen=True)
class WorksheetInventory:
    """Structural inventory for one worksheet."""

    position: int
    name: str
    max_row: int
    max_column: int
    populated_cell_count: int
    logical_data_row_count: int | None
    header_row: int | None
    headers: tuple[HeaderCellInventory, ...]
    tables: tuple[TableInventory, ...]
    formulas: tuple[FormulaInventory, ...]


@dataclass(frozen=True)
class WorkbookInventory:
    """File identity and complete structural inventory for one workbook."""

    filename: str
    workbook_type: str
    schema_version: str
    workbook_revision: str
    size_bytes: int
    sha256: str
    worksheet_count: int
    worksheets: tuple[WorksheetInventory, ...]
    populated_cell_count: int
    formula_count: int
    warning_count: int


@dataclass(frozen=True)
class IncomingPairInventory:
    """A fully inventoried incoming pair that has not yet been published."""

    incoming_pair: IncomingWorkbookPair
    glossary_inventory: WorkbookInventory
    references_inventory: WorkbookInventory
    data_release_id: str


@dataclass(frozen=True)
class RawSnapshotPublication:
    """A verified workbook pair published under one versioned directory."""

    inventory: IncomingPairInventory
    data_release_id: str
    snapshot_dir: Path
    glossary_path: Path
    references_path: Path


@dataclass(frozen=True)
class GitSourceState:
    """Local Git provenance recorded without workstation-specific details."""

    state: str
    head_commit: str | None


@dataclass(frozen=True)
class ManifestWorkbook:
    """One inventoried workbook and its portable raw-snapshot path."""

    inventory: WorkbookInventory
    snapshot_path: str


@dataclass(frozen=True)
class IntakeManifest:
    """Deterministic machine-readable identity of one raw snapshot pair."""

    manifest_schema_version: str
    data_release_id: str
    source_git: GitSourceState
    workbooks: tuple[ManifestWorkbook, ...]


@dataclass(frozen=True)
class IntakeManifestPublication:
    """An atomically published intake manifest and its exact bytes."""

    raw_snapshot: RawSnapshotPublication
    manifest: IntakeManifest
    path: Path
    serialized_bytes: bytes
    sha256: str


@dataclass(frozen=True)
class IntakePublication:
    """A complete raw snapshot and manifest, newly made or safely reused."""

    inventory: IncomingPairInventory
    raw_snapshot: RawSnapshotPublication
    manifest: IntakeManifestPublication
    disposition: str
    raw_created: bool
    manifest_created: bool


@dataclass(frozen=True)
class _SheetInventorySpec:
    """Internal structural role for one supported worksheet."""

    name: str
    table_name: str | None = None
    header_row: int | None = None
    require_content: bool = True


_GLOSSARY_SPECS = (
    _SheetInventorySpec(INTRODUCTION_SHEET_NAME),
    _SheetInventorySpec(GLOSSARY_SHEET_NAME, table_name=GLOSSARY_TABLE_NAME),
    _SheetInventorySpec(FOOTNOTES_SHEET_NAME, header_row=FOOTNOTES_HEADER_ROW),
    _SheetInventorySpec(METADATA_SHEET_NAME, table_name=METADATA_TABLE_NAME),
)
_REFERENCES_SPECS = (
    _SheetInventorySpec(REFERENCES_SHEET_NAME, header_row=REFERENCES_HEADER_ROW),
    _SheetInventorySpec(METADATA_SHEET_NAME, table_name=METADATA_TABLE_NAME),
)
_SHA256_PATTERN = re.compile(r"^[0-9a-f]{64}$")


def _read_required_workbook(path: Path, role: str) -> WorkbookSnapshot:
    try:
        return read_workbook(path)
    except ExcelReadError as error:
        raise IncomingContractError(
            f"could not read required {role} workbook at {path}: {error}"
        ) from error


def _extract_required_metadata(
    snapshot: WorkbookSnapshot,
    role: str,
) -> WorkbookMetadata:
    try:
        return extract_workbook_metadata(snapshot)
    except ValueError as error:
        raise IncomingContractError(
            f"invalid {role} workbook Metadata at {snapshot.path}: {error}"
        ) from error


def read_incoming_pair(
    incoming_dir: str | Path = INCOMING_DIR,
) -> IncomingWorkbookPair:
    """Read the stable incoming filenames and validate their Metadata pair."""

    directory = Path(incoming_dir).resolve()
    if not directory.is_dir():
        raise IncomingContractError(
            f"incoming directory does not exist or is not a directory: {directory}"
        )

    glossary_path = directory / GLOSSARY_WORKBOOK_FILENAME
    references_path = directory / REFERENCES_WORKBOOK_FILENAME
    glossary_snapshot = _read_required_workbook(glossary_path, "glossary")
    references_snapshot = _read_required_workbook(references_path, "references")

    glossary_metadata = _extract_required_metadata(
        glossary_snapshot,
        "glossary",
    )
    references_metadata = _extract_required_metadata(
        references_snapshot,
        "references",
    )
    try:
        compatibility = validate_workbook_compatibility(
            glossary_metadata,
            references_metadata,
        )
    except ValueError as error:
        raise IncomingContractError(
            f"incoming workbook pair is incompatible: {error}"
        ) from error

    return IncomingWorkbookPair(
        glossary_snapshot=glossary_snapshot,
        references_snapshot=references_snapshot,
        compatibility=compatibility,
    )


def _cell_map(
    sheet: WorksheetSnapshot,
    workbook_role: str,
) -> dict[tuple[int, int], CellSnapshot]:
    cells: dict[tuple[int, int], CellSnapshot] = {}
    for cell in sheet.cells:
        try:
            position = coordinate_to_tuple(cell.coordinate)
        except (TypeError, ValueError) as error:
            raise IncomingContractError(
                f"{workbook_role} workbook sheet {sheet.name!r} has invalid "
                f"cell coordinate {cell.coordinate!r}"
            ) from error
        if position in cells:
            raise IncomingContractError(
                f"{workbook_role} workbook sheet {sheet.name!r} has duplicate "
                f"cell coordinate {cell.coordinate!r}"
            )
        cells[position] = cell
    return cells


def _header_inventory(
    *,
    cells: dict[tuple[int, int], CellSnapshot],
    row: int,
    columns: tuple[int, ...],
    workbook_role: str,
    sheet_name: str,
    table_name: str | None = None,
) -> tuple[HeaderCellInventory, ...]:
    headers = []
    seen: set[str] = set()
    context = f" sheet {sheet_name!r}"
    if table_name is not None:
        context += f" table {table_name!r}"

    for column in columns:
        cell = cells.get((row, column))
        coordinate = cell.coordinate if cell is not None else f"column {column}"
        value = cell.value if cell is not None else None
        if (
            cell is None
            or cell.formula is not None
            or not isinstance(value, str)
            or not value.strip()
        ):
            raise IncomingContractError(
                f"{workbook_role} workbook{context} header at {coordinate} "
                "must be nonblank literal text"
            )
        if value in seen:
            raise IncomingContractError(
                f"{workbook_role} workbook{context} has duplicate header "
                f"{value!r} at {coordinate}"
            )
        seen.add(value)
        headers.append(HeaderCellInventory(column, cell.coordinate, value))
    return tuple(headers)


def _formula_inventory(
    sheet: WorksheetSnapshot,
    workbook_role: str,
) -> tuple[FormulaInventory, ...]:
    formulas = []
    for cell in sheet.cells:
        if cell.formula is None:
            continue
        if not isinstance(cell.formula, str) or not cell.formula:
            raise IncomingContractError(
                f"{workbook_role} workbook sheet {sheet.name!r} formula at "
                f"{cell.coordinate!r} must be nonblank text"
            )
        formulas.append(
            FormulaInventory(
                worksheet=sheet.name,
                coordinate=cell.coordinate,
                definition=cell.formula,
                has_cached_value=cell.value is not None,
            )
        )
    try:
        return tuple(
            sorted(formulas, key=lambda formula: coordinate_to_tuple(formula.coordinate))
        )
    except (TypeError, ValueError) as error:
        raise IncomingContractError(
            f"{workbook_role} workbook sheet {sheet.name!r} has an invalid "
            "formula coordinate"
        ) from error


def _table_inventory(
    *,
    table: TableSnapshot,
    sheet: WorksheetSnapshot,
    cells: dict[tuple[int, int], CellSnapshot],
    workbook_role: str,
) -> TableInventory:
    try:
        min_column, min_row, max_column, max_row = range_boundaries(
            table.reference
        )
    except (TypeError, ValueError) as error:
        raise IncomingContractError(
            f"{workbook_role} workbook sheet {sheet.name!r} table "
            f"{table.name!r} has malformed range {table.reference!r}"
        ) from error

    if (
        min_row < 1
        or min_column < 1
        or max_row > sheet.max_row
        or max_column > sheet.max_column
        or max_row <= min_row
    ):
        raise IncomingContractError(
            f"{workbook_role} workbook sheet {sheet.name!r} table "
            f"{table.name!r} has impossible range {table.reference!r}"
        )

    columns = tuple(range(min_column, max_column + 1))
    headers = _header_inventory(
        cells=cells,
        row=min_row,
        columns=columns,
        workbook_role=workbook_role,
        sheet_name=sheet.name,
        table_name=table.name,
    )
    populated_rows = {
        row
        for row, column in cells
        if min_row < row <= max_row and min_column <= column <= max_column
    }
    declared_rows = set(range(min_row + 1, max_row + 1))
    missing_rows = sorted(declared_rows - populated_rows)
    if missing_rows:
        raise IncomingContractError(
            f"{workbook_role} workbook sheet {sheet.name!r} table "
            f"{table.name!r} has blank declared data row {missing_rows[0]}"
        )

    outside = [
        cell.coordinate
        for position, cell in cells.items()
        if not (
            min_row <= position[0] <= max_row
            and min_column <= position[1] <= max_column
        )
    ]
    if outside:
        first = sorted(outside, key=coordinate_to_tuple)[0]
        raise IncomingContractError(
            f"{workbook_role} workbook sheet {sheet.name!r} has populated "
            f"cell {first} outside required table {table.name!r} range "
            f"{table.reference!r}"
        )

    return TableInventory(
        name=table.name,
        reference=table.reference,
        min_row=min_row,
        max_row=max_row,
        min_column=min_column,
        max_column=max_column,
        header_row=min_row,
        headers=headers,
        declared_data_row_count=max_row - min_row,
        populated_data_row_count=len(populated_rows),
    )


def _worksheet_inventory(
    *,
    sheet: WorksheetSnapshot,
    position: int,
    spec: _SheetInventorySpec,
    workbook_role: str,
) -> WorksheetInventory:
    cells = _cell_map(sheet, workbook_role)
    formulas = _formula_inventory(sheet, workbook_role)
    table_names = [table.name for table in sheet.tables]
    duplicate_tables = sorted(
        name for name, count in Counter(table_names).items() if count > 1
    )
    if duplicate_tables:
        raise IncomingContractError(
            f"{workbook_role} workbook sheet {sheet.name!r} has duplicate "
            f"table {duplicate_tables[0]!r}"
        )

    expected_tables = () if spec.table_name is None else (spec.table_name,)
    if set(table_names) != set(expected_tables) or len(table_names) != len(
        expected_tables
    ):
        raise IncomingContractError(
            f"{workbook_role} workbook sheet {sheet.name!r} expected tables "
            f"{expected_tables!r}; found {tuple(sorted(table_names))!r}"
        )

    tables = tuple(
        _table_inventory(
            table=table,
            sheet=sheet,
            cells=cells,
            workbook_role=workbook_role,
        )
        for table in sorted(sheet.tables, key=lambda item: (item.name, item.reference))
    )

    header_row = None
    headers: tuple[HeaderCellInventory, ...] = ()
    logical_rows = None
    if tables:
        table = tables[0]
        header_row = table.header_row
        headers = table.headers
        logical_rows = table.populated_data_row_count
    elif spec.header_row is not None:
        header_row = spec.header_row
        header_columns = tuple(
            sorted(column for row, column in cells if row == header_row)
        )
        if not header_columns:
            raise IncomingContractError(
                f"{workbook_role} workbook sheet {sheet.name!r} has no "
                f"headers at configured row {header_row}"
            )
        headers = _header_inventory(
            cells=cells,
            row=header_row,
            columns=header_columns,
            workbook_role=workbook_role,
            sheet_name=sheet.name,
        )
        header_column_set = set(header_columns)
        populated_rows = {
            row
            for row, column in cells
            if row > header_row and column in header_column_set
        }
        outside = [
            cell.coordinate
            for (row, column), cell in cells.items()
            if row > header_row and column not in header_column_set
        ]
        if outside:
            first = sorted(outside, key=coordinate_to_tuple)[0]
            raise IncomingContractError(
                f"{workbook_role} workbook sheet {sheet.name!r} has populated "
                f"cell {first} without a header"
            )
        if not populated_rows:
            raise IncomingContractError(
                f"{workbook_role} workbook sheet {sheet.name!r} has no "
                "populated data rows"
            )
        logical_rows = len(populated_rows)
    elif spec.require_content and not cells:
        raise IncomingContractError(
            f"{workbook_role} workbook sheet {sheet.name!r} has no content"
        )

    return WorksheetInventory(
        position=position,
        name=sheet.name,
        max_row=sheet.max_row,
        max_column=sheet.max_column,
        populated_cell_count=len(cells),
        logical_data_row_count=logical_rows,
        header_row=header_row,
        headers=headers,
        tables=tables,
        formulas=formulas,
    )


def _workbook_inventory(
    *,
    snapshot: WorkbookSnapshot,
    metadata: WorkbookMetadata,
    workbook_role: str,
    expected_filename: str,
    expected_type: str,
    expected_sheet_names: tuple[str, ...],
    specs: tuple[_SheetInventorySpec, ...],
) -> WorkbookInventory:
    if snapshot.path.name != expected_filename:
        raise IncomingContractError(
            f"{workbook_role} workbook must use filename {expected_filename!r}; "
            f"found {snapshot.path.name!r}"
        )
    if metadata.workbook_type != expected_type:
        raise IncomingContractError(
            f"{workbook_role} workbook must declare type {expected_type!r}"
        )
    if (
        not isinstance(snapshot.size_bytes, int)
        or isinstance(snapshot.size_bytes, bool)
        or snapshot.size_bytes <= 0
    ):
        raise IncomingContractError(
            f"{workbook_role} workbook {snapshot.path} has missing or invalid "
            "source byte size"
        )
    if not isinstance(snapshot.sha256, str) or not _SHA256_PATTERN.fullmatch(
        snapshot.sha256
    ):
        raise IncomingContractError(
            f"{workbook_role} workbook {snapshot.path} has missing or invalid "
            "SHA-256 digest"
        )

    sheet_names = [sheet.name for sheet in snapshot.sheets]
    counts = Counter(sheet_names)
    duplicates = sorted(name for name, count in counts.items() if count > 1)
    missing = sorted(set(expected_sheet_names) - set(sheet_names))
    unknown = sorted(set(sheet_names) - set(expected_sheet_names))
    if duplicates or missing or unknown:
        details = []
        if missing:
            details.append(f"missing sheets {tuple(missing)!r}")
        if duplicates:
            details.append(f"duplicate sheets {tuple(duplicates)!r}")
        if unknown:
            details.append(f"unknown sheets {tuple(unknown)!r}")
        raise IncomingContractError(
            f"{workbook_role} workbook structure is incomplete: "
            + "; ".join(details)
        )

    specs_by_name = {spec.name: spec for spec in specs}
    worksheets = tuple(
        _worksheet_inventory(
            sheet=sheet,
            position=position,
            spec=specs_by_name[sheet.name],
            workbook_role=workbook_role,
        )
        for position, sheet in enumerate(snapshot.sheets, start=1)
    )
    return WorkbookInventory(
        filename=snapshot.path.name,
        workbook_type=metadata.workbook_type,
        schema_version=metadata.schema_version,
        workbook_revision=metadata.workbook_revision,
        size_bytes=snapshot.size_bytes,
        sha256=snapshot.sha256,
        worksheet_count=len(worksheets),
        worksheets=worksheets,
        populated_cell_count=sum(
            sheet.populated_cell_count for sheet in worksheets
        ),
        formula_count=sum(len(sheet.formulas) for sheet in worksheets),
        warning_count=len(snapshot.warnings),
    )


def inventory_incoming_pair(pair: IncomingWorkbookPair) -> IncomingPairInventory:
    """Return a deterministic structural inventory of an accepted pair."""

    if not isinstance(pair, IncomingWorkbookPair):
        raise IncomingContractError(
            "workbook inventory requires an IncomingWorkbookPair"
        )
    compatibility = pair.compatibility
    glossary = _workbook_inventory(
        snapshot=pair.glossary_snapshot,
        metadata=compatibility.glossary_metadata,
        workbook_role="glossary",
        expected_filename=GLOSSARY_WORKBOOK_FILENAME,
        expected_type=GLOSSARY_WORKBOOK_TYPE,
        expected_sheet_names=GLOSSARY_WORKSHEET_NAMES,
        specs=_GLOSSARY_SPECS,
    )
    references = _workbook_inventory(
        snapshot=pair.references_snapshot,
        metadata=compatibility.references_metadata,
        workbook_role="references",
        expected_filename=REFERENCES_WORKBOOK_FILENAME,
        expected_type=REFERENCES_WORKBOOK_TYPE,
        expected_sheet_names=REFERENCES_WORKSHEET_NAMES,
        specs=_REFERENCES_SPECS,
    )
    return IncomingPairInventory(
        incoming_pair=pair,
        glossary_inventory=glossary,
        references_inventory=references,
        data_release_id=compatibility.data_release_id,
    )


def _publication_sources(inventory: IncomingPairInventory):
    pair = inventory.incoming_pair
    return (
        (
            "glossary",
            pair.glossary_snapshot,
            inventory.glossary_inventory,
            pair.compatibility.glossary_metadata,
            GLOSSARY_WORKBOOK_FILENAME,
        ),
        (
            "references",
            pair.references_snapshot,
            inventory.references_inventory,
            pair.compatibility.references_metadata,
            REFERENCES_WORKBOOK_FILENAME,
        ),
    )


def _validate_publication_inventory(inventory: object) -> IncomingPairInventory:
    if not isinstance(inventory, IncomingPairInventory):
        raise IncomingContractError(
            "raw snapshot publication requires an IncomingPairInventory"
        )
    release_id = validate_release_id(inventory.data_release_id)
    compatibility = inventory.incoming_pair.compatibility
    if release_id != compatibility.data_release_id:
        raise IncomingContractError(
            "raw snapshot inventory release ID does not match its incoming pair"
        )

    for role, snapshot, workbook, metadata, expected_filename in (
        _publication_sources(inventory)
    ):
        if snapshot.path.name != expected_filename or workbook.filename != (
            expected_filename
        ):
            raise IncomingContractError(
                f"raw snapshot {role} filename does not match "
                f"{expected_filename!r}"
            )
        if workbook.size_bytes != snapshot.size_bytes:
            raise IncomingContractError(
                f"raw snapshot {role} byte size does not match its inventory"
            )
        if workbook.sha256 != snapshot.sha256:
            raise IncomingContractError(
                f"raw snapshot {role} SHA-256 does not match its inventory"
            )
        if (
            workbook.workbook_type != metadata.workbook_type
            or workbook.schema_version != metadata.schema_version
            or workbook.workbook_revision != metadata.workbook_revision
        ):
            raise IncomingContractError(
                f"raw snapshot {role} Metadata does not match its inventory"
            )
        if snapshot.path.is_symlink() or not snapshot.path.is_file():
            raise IncomingContractError(
                f"raw snapshot {role} source is not an ordinary file: "
                f"{snapshot.path}"
            )
    return inventory


def _file_fingerprint(path: Path) -> tuple[int, str]:
    content = path.read_bytes()
    return len(content), sha256(content).hexdigest()


def publish_raw_snapshot(
    inventory: IncomingPairInventory,
    raw_snapshots_root: str | Path = RAW_SNAPSHOTS_DIR,
) -> RawSnapshotPublication:
    """Atomically publish a verified workbook pair under its release ID."""

    accepted = _validate_publication_inventory(inventory)
    release_id = accepted.data_release_id
    supplied_root = Path(raw_snapshots_root)
    if supplied_root.is_symlink():
        raise IncomingContractError(
            f"raw snapshot root must not be a symlink: {supplied_root}"
        )
    root = supplied_root.resolve()
    if root.exists() and not root.is_dir():
        raise IncomingContractError(
            f"raw snapshot root is not a directory: {root}"
        )
    target = root / validate_release_id(release_id)
    if target.exists() or target.is_symlink():
        raise IncomingContractError(
            f"raw snapshot target already exists; refusing to overwrite: {target}"
        )

    try:
        root.mkdir(parents=True, exist_ok=True)
        staging = Path(tempfile.mkdtemp(prefix=f".{release_id}-", dir=root))
    except OSError as error:
        raise IncomingContractError(
            f"raw snapshot {release_id} could not create staging under {root}"
        ) from error
    try:
        for role, snapshot, workbook, _metadata, filename in (
            _publication_sources(accepted)
        ):
            staged_path = staging / filename
            try:
                shutil.copyfile(snapshot.path, staged_path)
                staged_size, staged_hash = _file_fingerprint(staged_path)
            except OSError as error:
                raise IncomingContractError(
                    f"raw snapshot {release_id} failed to copy or verify "
                    f"{role} source {snapshot.path} to {staged_path}"
                ) from error
            if (
                staged_size != workbook.size_bytes
                or staged_hash != workbook.sha256
            ):
                raise IncomingContractError(
                    f"raw snapshot {release_id} staged {role} bytes do not "
                    "match the accepted inventory"
                )

        expected_names = {
            GLOSSARY_WORKBOOK_FILENAME,
            REFERENCES_WORKBOOK_FILENAME,
        }
        staged_entries = tuple(staging.iterdir())
        staged_names = {path.name for path in staged_entries}
        if staged_names != expected_names or any(
            path.is_symlink() or not path.is_file() for path in staged_entries
        ):
            raise IncomingContractError(
                f"raw snapshot {release_id} staging directory must contain "
                "exactly the two ordinary workbook files"
            )
        try:
            staging.replace(target)
        except OSError as error:
            raise IncomingContractError(
                f"raw snapshot {release_id} could not publish staging "
                f"directory {staging} to {target}"
            ) from error
    except Exception as error:
        if staging.exists():
            try:
                shutil.rmtree(staging)
            except OSError as cleanup_error:
                raise IncomingContractError(
                    f"raw snapshot {release_id} failed and staging cleanup "
                    f"also failed at {staging}: {cleanup_error}"
                ) from error
        raise

    return RawSnapshotPublication(
        inventory=accepted,
        data_release_id=release_id,
        snapshot_dir=target,
        glossary_path=target / GLOSSARY_WORKBOOK_FILENAME,
        references_path=target / REFERENCES_WORKBOOK_FILENAME,
    )


_GIT_COMMIT_PATTERN = re.compile(r"^(?:[0-9a-f]{40}|[0-9a-f]{64})$")
_GIT_STATES = frozenset({"commit", "local", "unknown"})


def validate_git_source_state(state: GitSourceState) -> GitSourceState:
    """Return a coherent Git provenance record or raise a contract error."""

    if not isinstance(state, GitSourceState) or state.state not in _GIT_STATES:
        raise IncomingContractError("invalid Git source state")
    if state.state == "unknown":
        if state.head_commit is not None:
            raise IncomingContractError(
                "unknown Git source state must not include a HEAD commit"
            )
        return state
    if (
        not isinstance(state.head_commit, str)
        or not _GIT_COMMIT_PATTERN.fullmatch(state.head_commit)
    ):
        raise IncomingContractError(
            f"{state.state} Git source state requires a lowercase HEAD commit"
        )
    return state


def inspect_git_source_state(
    repository_root: str | Path = REPOSITORY_ROOT,
    runner=subprocess.run,
) -> GitSourceState:
    """Inspect local Git state without failing intake when it is unavailable."""

    root = Path(repository_root).resolve()
    common = {
        "cwd": root,
        "capture_output": True,
        "text": True,
        "check": False,
        "timeout": 10,
    }
    try:
        head_result = runner(
            ["git", "rev-parse", "--verify", "HEAD"],
            **common,
        )
        if head_result.returncode != 0:
            return GitSourceState("unknown", None)
        head = head_result.stdout.strip()
        if not _GIT_COMMIT_PATTERN.fullmatch(head):
            return GitSourceState("unknown", None)
        status_result = runner(
            [
                "git",
                "status",
                "--porcelain=v1",
                "--untracked-files=all",
            ],
            **common,
        )
        if status_result.returncode != 0:
            return GitSourceState("unknown", None)
    except (OSError, subprocess.SubprocessError):
        return GitSourceState("unknown", None)

    state = "local" if status_result.stdout else "commit"
    return validate_git_source_state(GitSourceState(state, head))


def _validate_raw_snapshot_for_manifest(
    publication: object,
    *,
    verify_files: bool = True,
) -> RawSnapshotPublication:
    if not isinstance(publication, RawSnapshotPublication):
        raise IncomingContractError(
            "intake manifest requires a RawSnapshotPublication"
        )
    release_id = validate_release_id(publication.data_release_id)
    inventory = publication.inventory
    if (
        release_id != inventory.data_release_id
        or release_id != inventory.incoming_pair.compatibility.data_release_id
        or publication.snapshot_dir.name != release_id
    ):
        raise IncomingContractError(
            "intake manifest raw snapshot release relationships are inconsistent"
        )
    expected_paths = (
        (
            "glossary",
            publication.glossary_path,
            inventory.glossary_inventory,
            GLOSSARY_WORKBOOK_FILENAME,
        ),
        (
            "references",
            publication.references_path,
            inventory.references_inventory,
            REFERENCES_WORKBOOK_FILENAME,
        ),
    )
    for role, path, workbook, filename in expected_paths:
        if (
            path.parent != publication.snapshot_dir
            or path.name != filename
            or workbook.filename != filename
        ):
            raise IncomingContractError(
                f"intake manifest {role} snapshot path is inconsistent"
            )
    if not verify_files:
        return publication
    if publication.snapshot_dir.is_symlink() or not publication.snapshot_dir.is_dir():
        raise IncomingContractError(
            f"intake manifest raw snapshot directory is unavailable: "
            f"{publication.snapshot_dir}"
        )
    entries = tuple(publication.snapshot_dir.iterdir())
    expected_names = {GLOSSARY_WORKBOOK_FILENAME, REFERENCES_WORKBOOK_FILENAME}
    if {path.name for path in entries} != expected_names or any(
        path.is_symlink() or not path.is_file() for path in entries
    ):
        raise IncomingContractError(
            f"intake manifest raw snapshot {release_id} must contain exactly "
            "the two ordinary workbook files"
        )
    for role, path, workbook, _filename in expected_paths:
        try:
            size, digest = _file_fingerprint(path)
        except OSError as error:
            raise IncomingContractError(
                f"intake manifest could not verify {role} snapshot {path}"
            ) from error
        if size != workbook.size_bytes or digest != workbook.sha256:
            raise IncomingContractError(
                f"intake manifest {role} snapshot bytes do not match inventory"
            )
    return publication


def build_intake_manifest(
    publication: RawSnapshotPublication,
    source_git: GitSourceState,
) -> IntakeManifest:
    """Build a portable manifest after revalidating the published raw pair."""

    accepted = _validate_raw_snapshot_for_manifest(
        publication,
        verify_files=False,
    )
    git_state = validate_git_source_state(source_git)
    release_id = accepted.data_release_id
    inventory = accepted.inventory
    return IntakeManifest(
        manifest_schema_version=INTAKE_MANIFEST_SCHEMA_VERSION,
        data_release_id=release_id,
        source_git=git_state,
        workbooks=(
            ManifestWorkbook(
                inventory.glossary_inventory,
                f"{release_id}/{GLOSSARY_WORKBOOK_FILENAME}",
            ),
            ManifestWorkbook(
                inventory.references_inventory,
                f"{release_id}/{REFERENCES_WORKBOOK_FILENAME}",
            ),
        ),
    )


def _header_dict(header: HeaderCellInventory) -> dict[str, object]:
    return {
        "column": header.column,
        "coordinate": header.coordinate,
        "value": header.value,
    }


def _formula_dict(formula: FormulaInventory) -> dict[str, object]:
    return {
        "coordinate": formula.coordinate,
        "definition": formula.definition,
        "has_cached_value": formula.has_cached_value,
        "worksheet": formula.worksheet,
    }


def _table_dict(table: TableInventory) -> dict[str, object]:
    return {
        "declared_data_row_count": table.declared_data_row_count,
        "header_row": table.header_row,
        "headers": [_header_dict(header) for header in table.headers],
        "max_column": table.max_column,
        "max_row": table.max_row,
        "min_column": table.min_column,
        "min_row": table.min_row,
        "name": table.name,
        "populated_data_row_count": table.populated_data_row_count,
        "reference": table.reference,
    }


def _worksheet_dict(worksheet: WorksheetInventory) -> dict[str, object]:
    return {
        "formulas": [_formula_dict(formula) for formula in worksheet.formulas],
        "header_row": worksheet.header_row,
        "headers": [_header_dict(header) for header in worksheet.headers],
        "logical_data_row_count": worksheet.logical_data_row_count,
        "max_column": worksheet.max_column,
        "max_row": worksheet.max_row,
        "name": worksheet.name,
        "populated_cell_count": worksheet.populated_cell_count,
        "position": worksheet.position,
        "tables": [_table_dict(table) for table in worksheet.tables],
    }


def _manifest_dict(manifest: IntakeManifest) -> dict[str, object]:
    return {
        "data_release_id": manifest.data_release_id,
        "manifest_schema_version": manifest.manifest_schema_version,
        "source_git": {
            "head_commit": manifest.source_git.head_commit,
            "state": manifest.source_git.state,
        },
        "workbooks": [
            {
                "filename": workbook.inventory.filename,
                "formula_count": workbook.inventory.formula_count,
                "populated_cell_count": workbook.inventory.populated_cell_count,
                "schema_version": workbook.inventory.schema_version,
                "sha256": workbook.inventory.sha256,
                "size_bytes": workbook.inventory.size_bytes,
                "snapshot_path": workbook.snapshot_path,
                "warning_count": workbook.inventory.warning_count,
                "workbook_revision": workbook.inventory.workbook_revision,
                "workbook_type": workbook.inventory.workbook_type,
                "worksheet_count": workbook.inventory.worksheet_count,
                "worksheets": [
                    _worksheet_dict(worksheet)
                    for worksheet in workbook.inventory.worksheets
                ],
            }
            for workbook in manifest.workbooks
        ],
    }


def serialize_intake_manifest(manifest: IntakeManifest) -> bytes:
    """Return deterministic UTF-8 JSON for a validated intake manifest."""

    if not isinstance(manifest, IntakeManifest):
        raise IncomingContractError("expected an IntakeManifest")
    validate_release_id(manifest.data_release_id)
    if manifest.manifest_schema_version != INTAKE_MANIFEST_SCHEMA_VERSION:
        raise IncomingContractError("unsupported intake manifest schema version")
    validate_git_source_state(manifest.source_git)
    if len(manifest.workbooks) != 2:
        raise IncomingContractError("intake manifest must contain two workbooks")
    text = json.dumps(
        _manifest_dict(manifest),
        ensure_ascii=False,
        indent=2,
        sort_keys=True,
    )
    return (text + "\n").encode("utf-8")


def publish_intake_manifest(
    publication: RawSnapshotPublication,
    manifest_root: str | Path = MANIFEST_DIR,
    *,
    source_git: GitSourceState | None = None,
    repository_root: str | Path = REPOSITORY_ROOT,
    git_runner=subprocess.run,
) -> IntakeManifestPublication:
    """Atomically publish deterministic JSON for one verified raw snapshot."""

    accepted = _validate_raw_snapshot_for_manifest(publication)
    supplied_root = Path(manifest_root)
    if supplied_root.is_symlink():
        raise IncomingContractError(
            f"intake manifest root must not be a symlink: {supplied_root}"
        )
    root = supplied_root.resolve()
    if root.exists() and not root.is_dir():
        raise IncomingContractError(
            f"intake manifest root is not a directory: {root}"
        )
    target = root / f"{validate_release_id(accepted.data_release_id)}.json"
    if target.exists() or target.is_symlink():
        raise IncomingContractError(
            f"intake manifest target already exists; refusing to overwrite: {target}"
        )

    git_state = source_git or inspect_git_source_state(
        repository_root,
        runner=git_runner,
    )
    manifest = build_intake_manifest(accepted, git_state)
    serialized = serialize_intake_manifest(manifest)
    try:
        root.mkdir(parents=True, exist_ok=True)
    except OSError as error:
        raise IncomingContractError(
            f"intake manifest {accepted.data_release_id} could not create "
            f"its root directory {root}"
        ) from error
    file_descriptor = None
    temporary_name = None
    try:
        file_descriptor, temporary_name = tempfile.mkstemp(
            prefix=f".{accepted.data_release_id}-",
            suffix=".json",
            dir=root,
        )
        os.close(file_descriptor)
    except OSError as error:
        if file_descriptor is not None:
            try:
                os.close(file_descriptor)
            except OSError:
                pass
        if temporary_name is not None and Path(temporary_name).exists():
            try:
                Path(temporary_name).unlink()
            except OSError as cleanup_error:
                raise IncomingContractError(
                    f"intake manifest {accepted.data_release_id} temporary "
                    f"creation failed and cleanup also failed at "
                    f"{temporary_name}: {cleanup_error}"
                ) from error
        raise IncomingContractError(
            f"intake manifest {accepted.data_release_id} could not create "
            f"a temporary file under {root}"
        ) from error
    temporary_path = Path(temporary_name)
    try:
        try:
            temporary_path.write_bytes(serialized)
            if temporary_path.read_bytes() != serialized:
                raise IncomingContractError(
                    f"intake manifest temporary verification failed: "
                    f"{temporary_path}"
                )
            temporary_path.replace(target)
        except OSError as error:
            raise IncomingContractError(
                f"intake manifest {accepted.data_release_id} could not publish "
                f"temporary file {temporary_path} to {target}"
            ) from error
    except Exception as error:
        if temporary_path.exists():
            try:
                temporary_path.unlink()
            except OSError as cleanup_error:
                raise IncomingContractError(
                    f"intake manifest {accepted.data_release_id} failed and "
                    f"temporary cleanup also failed at {temporary_path}: "
                    f"{cleanup_error}"
                ) from error
        raise

    return IntakeManifestPublication(
        raw_snapshot=accepted,
        manifest=manifest,
        path=target,
        serialized_bytes=serialized,
        sha256=sha256(serialized).hexdigest(),
    )


_INTAKE_DISPOSITIONS = frozenset({"created", "recovered", "existing"})
_MANIFEST_TOP_LEVEL_KEYS = frozenset(
    {"data_release_id", "manifest_schema_version", "source_git", "workbooks"}
)
_MANIFEST_WORKBOOK_KEYS = frozenset(
    {
        "filename",
        "formula_count",
        "populated_cell_count",
        "schema_version",
        "sha256",
        "size_bytes",
        "snapshot_path",
        "warning_count",
        "workbook_revision",
        "workbook_type",
        "worksheet_count",
        "worksheets",
    }
)
_MANIFEST_WORKSHEET_KEYS = frozenset(
    {
        "formulas",
        "header_row",
        "headers",
        "logical_data_row_count",
        "max_column",
        "max_row",
        "name",
        "populated_cell_count",
        "position",
        "tables",
    }
)
_MANIFEST_HEADER_KEYS = frozenset({"column", "coordinate", "value"})
_MANIFEST_FORMULA_KEYS = frozenset(
    {"coordinate", "definition", "has_cached_value", "worksheet"}
)
_MANIFEST_TABLE_KEYS = frozenset(
    {
        "declared_data_row_count",
        "header_row",
        "headers",
        "max_column",
        "max_row",
        "min_column",
        "min_row",
        "name",
        "populated_data_row_count",
        "reference",
    }
)
_MANIFEST_TEMPORARY_PATTERN = re.compile(
    r"^\.[0-9]{8}(?:-r(?:[2-9]|[1-9][0-9]+))?-.+\.json$"
)


def _validated_artifact_root(path: str | Path, label: str) -> Path:
    supplied = Path(path)
    if supplied.is_symlink():
        raise IncomingContractError(f"{label} root must not be a symlink: {supplied}")
    root = supplied.resolve()
    if root.exists() and not root.is_dir():
        raise IncomingContractError(f"{label} root is not a directory: {root}")
    return root


def _existing_raw_snapshot(
    inventory: IncomingPairInventory,
    raw_root: Path,
) -> RawSnapshotPublication | None:
    release_id = inventory.data_release_id
    target = raw_root / validate_release_id(release_id)
    if target.is_symlink():
        raise IncomingContractError(
            f"raw snapshot collision for {release_id}: target is a symlink"
        )
    if not target.exists():
        return None
    if not target.is_dir():
        raise IncomingContractError(
            f"raw snapshot collision for {release_id}: target is not a directory"
        )

    expected = {
        GLOSSARY_WORKBOOK_FILENAME: inventory.glossary_inventory,
        REFERENCES_WORKBOOK_FILENAME: inventory.references_inventory,
    }
    try:
        entries = tuple(target.iterdir())
    except OSError as error:
        raise IncomingContractError(
            f"raw snapshot collision for {release_id}: target is unreadable"
        ) from error
    if {entry.name for entry in entries} != set(expected) or any(
        entry.is_symlink() or not entry.is_file() for entry in entries
    ):
        raise IncomingContractError(
            f"raw snapshot collision for {release_id}: expected exactly the "
            "two ordinary workbook files"
        )
    for filename, workbook in expected.items():
        path = target / filename
        try:
            size, digest = _file_fingerprint(path)
        except OSError as error:
            raise IncomingContractError(
                f"raw snapshot collision for {release_id}: {filename} is unreadable"
            ) from error
        if size != workbook.size_bytes:
            raise IncomingContractError(
                f"raw snapshot collision for {release_id}: {workbook.workbook_type} "
                f"revision {workbook.workbook_revision} has a different size"
            )
        if digest != workbook.sha256:
            raise IncomingContractError(
                f"raw snapshot collision for {release_id}: {workbook.workbook_type} "
                f"revision {workbook.workbook_revision} has a different SHA-256"
            )

    return RawSnapshotPublication(
        inventory=inventory,
        data_release_id=release_id,
        snapshot_dir=target,
        glossary_path=target / GLOSSARY_WORKBOOK_FILENAME,
        references_path=target / REFERENCES_WORKBOOK_FILENAME,
    )


def _reject_duplicate_json_pairs(pairs):
    result = {}
    for key, value in pairs:
        if key in result:
            raise IncomingContractError(f"duplicate manifest JSON key: {key}")
        result[key] = value
    return result


def _reject_json_constant(value: str):
    raise IncomingContractError(f"invalid manifest JSON constant: {value}")


def _require_json_object(value: object, keys: frozenset[str], label: str) -> dict:
    if not isinstance(value, dict) or frozenset(value) != keys:
        raise IncomingContractError(f"invalid intake manifest {label}")
    return value


def _validate_manifest_structure(data: object, release_id: str) -> dict:
    document = _require_json_object(data, _MANIFEST_TOP_LEVEL_KEYS, "document")
    if document["manifest_schema_version"] != INTAKE_MANIFEST_SCHEMA_VERSION:
        raise IncomingContractError("unsupported intake manifest schema version")
    if document["data_release_id"] != release_id:
        raise IncomingContractError(
            f"intake manifest filename/content release mismatch for {release_id}"
        )
    source_git = _require_json_object(
        document["source_git"],
        frozenset({"head_commit", "state"}),
        "source_git",
    )
    validate_git_source_state(
        GitSourceState(source_git["state"], source_git["head_commit"])
    )

    workbooks = document["workbooks"]
    if not isinstance(workbooks, list) or len(workbooks) != 2:
        raise IncomingContractError("intake manifest must contain two workbooks")
    expected_roles = (
        (GLOSSARY_WORKBOOK_TYPE, GLOSSARY_WORKBOOK_FILENAME),
        (REFERENCES_WORKBOOK_TYPE, REFERENCES_WORKBOOK_FILENAME),
    )
    for workbook, (workbook_type, filename) in zip(workbooks, expected_roles):
        item = _require_json_object(
            workbook,
            _MANIFEST_WORKBOOK_KEYS,
            "workbook",
        )
        if (
            item["workbook_type"] != workbook_type
            or item["filename"] != filename
            or item["snapshot_path"] != f"{release_id}/{filename}"
        ):
            raise IncomingContractError(
                f"invalid intake manifest {workbook_type} identity or path"
            )
        if (
            not isinstance(item["workbook_revision"], str)
            or not item["workbook_revision"]
            or type(item["size_bytes"]) is not int
            or item["size_bytes"] < 0
            or not isinstance(item["sha256"], str)
            or not _SHA256_PATTERN.fullmatch(item["sha256"])
        ):
            raise IncomingContractError(
                f"invalid intake manifest {workbook_type} fingerprint"
            )
        worksheets = item["worksheets"]
        if not isinstance(worksheets, list):
            raise IncomingContractError("invalid intake manifest worksheets")
        for worksheet in worksheets:
            sheet = _require_json_object(
                worksheet,
                _MANIFEST_WORKSHEET_KEYS,
                "worksheet",
            )
            if not isinstance(sheet["headers"], list) or not isinstance(
                sheet["tables"], list
            ) or not isinstance(sheet["formulas"], list):
                raise IncomingContractError("invalid intake manifest worksheet arrays")
            for header in sheet["headers"]:
                _require_json_object(header, _MANIFEST_HEADER_KEYS, "header")
            for formula in sheet["formulas"]:
                _require_json_object(formula, _MANIFEST_FORMULA_KEYS, "formula")
            for table in sheet["tables"]:
                table_item = _require_json_object(
                    table,
                    _MANIFEST_TABLE_KEYS,
                    "table",
                )
                if not isinstance(table_item["headers"], list):
                    raise IncomingContractError("invalid intake manifest table headers")
                for header in table_item["headers"]:
                    _require_json_object(header, _MANIFEST_HEADER_KEYS, "header")
    return document


def _read_manifest_document(path: Path) -> tuple[dict, bytes]:
    if path.is_symlink() or not path.is_file():
        raise IncomingContractError(f"intake manifest is not an ordinary file: {path}")
    try:
        serialized = path.read_bytes()
        data = json.loads(
            serialized,
            object_pairs_hook=_reject_duplicate_json_pairs,
            parse_constant=_reject_json_constant,
        )
    except IncomingContractError:
        raise
    except (OSError, UnicodeError, json.JSONDecodeError) as error:
        raise IncomingContractError(f"could not read intake manifest: {path}") from error

    try:
        release_id = validate_release_id(path.stem)
    except ValueError as error:
        raise IncomingContractError(
            f"invalid intake manifest history filename: {path}"
        ) from error
    document = _validate_manifest_structure(data, release_id)
    canonical = (
        json.dumps(document, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    ).encode("utf-8")
    if canonical != serialized:
        raise IncomingContractError(f"intake manifest is not canonical: {path}")
    return document, serialized


def _existing_intake_manifest(
    raw_snapshot: RawSnapshotPublication,
    manifest_root: Path,
) -> IntakeManifestPublication | None:
    release_id = raw_snapshot.data_release_id
    target = manifest_root / f"{validate_release_id(release_id)}.json"
    if target.is_symlink():
        raise IncomingContractError(
            f"intake manifest collision for {release_id}: target is a symlink"
        )
    if not target.exists():
        return None
    if not target.is_file():
        raise IncomingContractError(
            f"intake manifest collision for {release_id}: target is not a file"
        )
    document, serialized = _read_manifest_document(target)
    source = document["source_git"]
    git_state = GitSourceState(source["state"], source["head_commit"])
    manifest = build_intake_manifest(raw_snapshot, git_state)
    expected = serialize_intake_manifest(manifest)
    if expected != serialized:
        raise IncomingContractError(
            f"intake manifest collision for {release_id}: existing inventory "
            "does not match the accepted source pair"
        )
    return IntakeManifestPublication(
        raw_snapshot=raw_snapshot,
        manifest=manifest,
        path=target,
        serialized_bytes=serialized,
        sha256=sha256(serialized).hexdigest(),
    )


def _check_manifest_revision_history(
    inventory: IncomingPairInventory,
    manifest_root: Path,
) -> None:
    if not manifest_root.exists():
        return
    current = {
        (workbook.workbook_type, workbook.workbook_revision): workbook
        for workbook in (
            inventory.glossary_inventory,
            inventory.references_inventory,
        )
    }
    try:
        entries = tuple(sorted(manifest_root.iterdir(), key=lambda path: path.name))
    except OSError as error:
        raise IncomingContractError(
            f"could not inspect intake manifest history: {manifest_root}"
        ) from error
    for path in entries:
        if _MANIFEST_TEMPORARY_PATTERN.fullmatch(path.name):
            continue
        if path.suffix != ".json":
            raise IncomingContractError(
                f"unexpected intake manifest history entry: {path}"
            )
        document, _serialized = _read_manifest_document(path)
        for historical in document["workbooks"]:
            identity = (
                historical["workbook_type"],
                historical["workbook_revision"],
            )
            workbook = current.get(identity)
            if workbook is None:
                continue
            if historical["size_bytes"] != workbook.size_bytes:
                raise IncomingContractError(
                    f"workbook revision collision: {workbook.workbook_type} "
                    f"revision {workbook.workbook_revision} has a different size"
                )
            if historical["sha256"] != workbook.sha256:
                raise IncomingContractError(
                    f"workbook revision collision: {workbook.workbook_type} "
                    f"revision {workbook.workbook_revision} has a different SHA-256"
                )


def publish_or_reuse_intake(
    inventory: IncomingPairInventory,
    raw_snapshots_root: str | Path = RAW_SNAPSHOTS_DIR,
    manifest_root: str | Path = MANIFEST_DIR,
    *,
    source_git: GitSourceState | None = None,
    repository_root: str | Path = REPOSITORY_ROOT,
    git_runner=subprocess.run,
) -> IntakePublication:
    """Publish one intake or safely reuse an exact completed attempt."""

    accepted = _validate_publication_inventory(inventory)
    raw_root = _validated_artifact_root(raw_snapshots_root, "raw snapshot")
    manifests = _validated_artifact_root(manifest_root, "intake manifest")
    _check_manifest_revision_history(accepted, manifests)

    raw_snapshot = _existing_raw_snapshot(accepted, raw_root)
    manifest_target = manifests / f"{accepted.data_release_id}.json"
    if raw_snapshot is None and (
        manifest_target.exists() or manifest_target.is_symlink()
    ):
        raise IncomingContractError(
            f"incomplete intake {accepted.data_release_id}: manifest exists "
            "without its raw snapshot"
        )

    raw_created = False
    if raw_snapshot is None:
        try:
            raw_snapshot = publish_raw_snapshot(accepted, raw_root)
            raw_created = True
        except IncomingContractError as publication_error:
            try:
                raw_snapshot = _existing_raw_snapshot(accepted, raw_root)
            except IncomingContractError:
                raise
            if raw_snapshot is None:
                raise publication_error

    existing_manifest = _existing_intake_manifest(raw_snapshot, manifests)
    manifest_created = False
    if existing_manifest is None:
        try:
            existing_manifest = publish_intake_manifest(
                raw_snapshot,
                manifests,
                source_git=source_git,
                repository_root=repository_root,
                git_runner=git_runner,
            )
            manifest_created = True
        except IncomingContractError as publication_error:
            try:
                existing_manifest = _existing_intake_manifest(
                    raw_snapshot,
                    manifests,
                )
            except IncomingContractError:
                raise
            if existing_manifest is None:
                raise publication_error

    final_raw = _existing_raw_snapshot(accepted, raw_root)
    if final_raw is None:
        raise IncomingContractError(
            f"completed intake {accepted.data_release_id} lost its raw snapshot"
        )
    final_manifest = _existing_intake_manifest(final_raw, manifests)
    if final_manifest is None:
        raise IncomingContractError(
            f"completed intake {accepted.data_release_id} lost its manifest"
        )

    if raw_created:
        disposition = "created"
    elif manifest_created:
        disposition = "recovered"
    else:
        disposition = "existing"
    if disposition not in _INTAKE_DISPOSITIONS:
        raise IncomingContractError("invalid intake publication disposition")
    return IntakePublication(
        inventory=accepted,
        raw_snapshot=final_raw,
        manifest=final_manifest,
        disposition=disposition,
        raw_created=raw_created,
        manifest_created=manifest_created,
    )
