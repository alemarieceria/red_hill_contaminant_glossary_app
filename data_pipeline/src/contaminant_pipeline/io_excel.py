"""Read-only boundary for Excel workbook contents and structure."""

from dataclasses import dataclass
from pathlib import Path
import warnings as python_warnings
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class ExcelReadWarning:
    """A non-blocking problem found while reading a workbook."""

    message: str
    sheet_name: str | None = None
    coordinate: str | None = None


@dataclass(frozen=True)
class CellSnapshot:
    """A nonblank or formula cell and its last Excel-calculated value."""

    coordinate: str
    value: object
    formula: str | None


@dataclass(frozen=True)
class TableSnapshot:
    """The name and cell range of an Excel table."""

    name: str
    reference: str


@dataclass(frozen=True)
class WorksheetSnapshot:
    """Read-only worksheet structure and populated cells."""

    name: str
    max_row: int
    max_column: int
    tables: tuple[TableSnapshot, ...]
    cells: tuple[CellSnapshot, ...]


@dataclass(frozen=True)
class WorkbookSnapshot:
    """A closed workbook represented entirely as ordinary Python data."""

    path: Path
    sheets: tuple[WorksheetSnapshot, ...]
    warnings: tuple[ExcelReadWarning, ...]


class ExcelReadError(ValueError):
    """Raised when an input is not a readable XLSX workbook."""


def _read_worksheet(
    formula_sheet: Worksheet,
    cached_sheet: Worksheet,
    read_warnings: list[ExcelReadWarning],
) -> WorksheetSnapshot:
    tables = tuple(
        TableSnapshot(name=table.name, reference=table.ref)
        for table in formula_sheet.tables.values()
    )
    cells: list[CellSnapshot] = []

    for row in formula_sheet.iter_rows():
        for formula_cell in row:
            cached_cell = cached_sheet[formula_cell.coordinate]
            formula = (
                str(formula_cell.value)
                if formula_cell.data_type == "f"
                else None
            )
            value = cached_cell.value if formula is not None else formula_cell.value

            if value is not None or formula is not None:
                cells.append(
                    CellSnapshot(
                        coordinate=formula_cell.coordinate,
                        value=value,
                        formula=formula,
                    )
                )

            comment = formula_cell.comment
            if comment is not None and (
                not isinstance(comment.author, str) or not comment.author.strip()
            ):
                read_warnings.append(
                    ExcelReadWarning(
                        message="comment has a blank or malformed author",
                        sheet_name=formula_sheet.title,
                        coordinate=formula_cell.coordinate,
                    )
                )

    return WorksheetSnapshot(
        name=formula_sheet.title,
        max_row=formula_sheet.max_row,
        max_column=formula_sheet.max_column,
        tables=tables,
        cells=tuple(cells),
    )


def read_workbook(path: str | Path) -> WorkbookSnapshot:
    """Return workbook contents without modifying or keeping the file open."""

    workbook_path = Path(path).resolve()
    if workbook_path.suffix.lower() != ".xlsx":
        raise ExcelReadError(f"expected an .xlsx workbook: {workbook_path}")
    if not workbook_path.is_file():
        raise ExcelReadError(f"workbook does not exist: {workbook_path}")

    formula_workbook = None
    cached_workbook = None
    read_warnings: list[ExcelReadWarning] = []

    try:
        with python_warnings.catch_warnings(record=True) as library_warnings:
            python_warnings.simplefilter("always")
            formula_workbook = load_workbook(
                workbook_path,
                data_only=False,
                read_only=False,
            )
            cached_workbook = load_workbook(
                workbook_path,
                data_only=True,
                read_only=False,
            )

        seen_messages: set[str] = set()
        for warning in library_warnings:
            message = str(warning.message)
            if message not in seen_messages:
                read_warnings.append(ExcelReadWarning(message=message))
                seen_messages.add(message)

        sheets = tuple(
            _read_worksheet(
                formula_sheet,
                cached_workbook[formula_sheet.title],
                read_warnings,
            )
            for formula_sheet in formula_workbook.worksheets
        )
    except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError) as error:
        raise ExcelReadError(f"could not read workbook: {workbook_path}") from error
    finally:
        if formula_workbook is not None:
            formula_workbook.close()
        if cached_workbook is not None:
            cached_workbook.close()

    return WorkbookSnapshot(
        path=workbook_path,
        sheets=sheets,
        warnings=tuple(read_warnings),
    )
