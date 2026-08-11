"""Build the Phase 3.3a scientific review handoff without editing sources."""

from dataclasses import dataclass, replace
from datetime import datetime
from enum import StrEnum
from pathlib import Path
import re
from types import MappingProxyType
from typing import Iterable, Mapping
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter, range_boundaries
from pydantic import HttpUrl, TypeAdapter, ValidationError

from .config import GLOSSARY_SHEET_NAME, GLOSSARY_TABLE_NAME
from .paths import supervisor_review_workbook_path
from .schemas import GLOSSARY_HEADER_MAP
from .scientific_validation import (
    ScientificFieldInspection,
    inspect_scientific_fields,
)
from .validate import (
    ValidatedIdentityRelationships,
    ValidationFinding,
    ValidationSeverity,
)


class ScientificReviewType(StrEnum):
    """Why one source field appears in the supervisor handoff."""

    INVALID_SCIENTIFIC_VALUE = "invalid_scientific_value"
    UNVERIFIED_NOT_APPLICABLE = "unverified_not_applicable"
    UNKNOWN_IDENTIFIER = "unknown_identifier"
    PENDING_SOURCE = "pending_source"


class ScientificReviewStatus(StrEnum):
    """Explicit review states; formatting is never the decision itself."""

    PROPOSED = "proposed"
    NEEDS_REVIEW = "needs_review"
    APPROVED_VALUE = "approved_value"
    APPROVED_NOT_APPLICABLE = "approved_not_applicable"
    UNKNOWN_PENDING = "unknown_pending"
    CORRECTION_REQUIRED = "correction_required"
    RESOLVED = "resolved"


class ReviewReconciliationStatus(StrEnum):
    """Relationship between an earlier review row and later findings."""

    RESOLVED = "resolved"
    STILL_FAILING = "still_failing"
    SUPERSEDED = "superseded"


_HTTP_URL = TypeAdapter(HttpUrl)
_ISO_DATE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
_IDENTIFIER_FIELDS = frozenset({"id_casrn", "id_inchikey"})
_SCIENTIFIC_REVIEW_FIELDS = frozenset(
    {"id_casrn", "id_inchikey", "id_chem_formula"}
)
_NOT_APPLICABLE_TOKENS = frozenset({"NA", "N/A"})
_REVIEW_STATUSES = tuple(status.value for status in ScientificReviewStatus)


def _required_text(value: object, label: str) -> str:
    if not isinstance(value, str) or not value or value != value.strip():
        raise ValueError(f"{label} must be nonblank text without surrounding whitespace")
    return value


def _optional_text(value: object, label: str) -> str | None:
    if value is None:
        return None
    return _required_text(value, label)


def _exact_nonempty_text(value: object, label: str) -> str:
    if not isinstance(value, str) or not value:
        raise ValueError(f"{label} must be nonempty text")
    return value


def _review_date(value: object, label: str) -> str | None:
    text = _optional_text(value, label)
    if text is None:
        return None
    if _ISO_DATE.fullmatch(text) is None:
        raise ValueError(f"{label} must use YYYY-MM-DD")
    datetime.strptime(text, "%Y-%m-%d")
    return text


def _evidence_url(value: object) -> str | None:
    text = _optional_text(value, "evidence_url")
    if text is None:
        return None
    try:
        return str(_HTTP_URL.validate_python(text))
    except ValidationError as error:
        raise ValueError("evidence_url must be an absolute HTTP(S) URL") from error


@dataclass(frozen=True)
class ScientificReviewProposal:
    """Independently supplied evidence for one proposed field decision."""

    id_contaminant: str
    canonical_field: str
    proposed_value: str | None = None
    source_system: str | None = None
    source_record_id: str | None = None
    evidence_url: str | None = None
    retrieval_date: str | None = None
    rationale: str | None = None
    status: ScientificReviewStatus = ScientificReviewStatus.PROPOSED
    reviewer: str | None = None
    review_date: str | None = None
    notes: str | None = None

    def __post_init__(self) -> None:
        object.__setattr__(
            self, "id_contaminant", _required_text(self.id_contaminant, "id_contaminant")
        )
        if re.fullmatch(r"RHC-\d{3}", self.id_contaminant) is None:
            raise ValueError("id_contaminant must use RHC-NNN")
        object.__setattr__(
            self, "canonical_field", _required_text(self.canonical_field, "canonical_field")
        )
        for field_name in (
            "proposed_value",
            "source_system",
            "source_record_id",
            "rationale",
            "reviewer",
            "notes",
        ):
            object.__setattr__(
                self,
                field_name,
                _optional_text(getattr(self, field_name), field_name),
            )
        object.__setattr__(self, "evidence_url", _evidence_url(self.evidence_url))
        object.__setattr__(
            self,
            "retrieval_date",
            _review_date(self.retrieval_date, "retrieval_date"),
        )
        object.__setattr__(
            self, "review_date", _review_date(self.review_date, "review_date")
        )
        if not isinstance(self.status, ScientificReviewStatus):
            raise TypeError("status must be ScientificReviewStatus")
        evidence = (self.source_system, self.source_record_id, self.evidence_url)
        if any(value is not None for value in evidence) and any(
            value is None for value in evidence
        ):
            raise ValueError(
                "evidence requires source_system, source_record_id, and evidence_url"
            )
        if self.proposed_value is not None and any(
            value is None for value in evidence
        ):
            raise ValueError(
                "a proposed value requires source_system, source_record_id, and evidence_url"
            )
        if self.status is ScientificReviewStatus.APPROVED_VALUE and (
            self.proposed_value is None or self.reviewer is None or self.review_date is None
        ):
            raise ValueError(
                "approved_value requires proposed_value, reviewer, and review_date"
            )
        if self.status is ScientificReviewStatus.APPROVED_NOT_APPLICABLE and (
            self.rationale is None or self.reviewer is None or self.review_date is None
        ):
            raise ValueError(
                "approved_not_applicable requires rationale, reviewer, and review_date"
            )
        if self.status is ScientificReviewStatus.PROPOSED and self.proposed_value is None:
            raise ValueError("proposed status requires proposed_value")
        if self.status is ScientificReviewStatus.RESOLVED:
            raise ValueError(
                "resolved is assigned by later reconciliation, not by a proposal"
            )


@dataclass(frozen=True)
class ResolvedCleanupRecord:
    """A separately supplied, already reviewed mechanical workbook correction."""

    id_contaminant: str
    id_name: str
    workbook: str
    sheet: str
    cell: str
    canonical_field: str
    old_value: str
    new_value: str
    correction_type: str
    notes: str | None = None

    def __post_init__(self) -> None:
        for field_name in (
            "id_contaminant",
            "id_name",
            "workbook",
            "sheet",
            "cell",
            "canonical_field",
            "correction_type",
        ):
            object.__setattr__(
                self,
                field_name,
                _required_text(getattr(self, field_name), field_name),
            )
        object.__setattr__(
            self, "old_value", _exact_nonempty_text(self.old_value, "old_value")
        )
        object.__setattr__(
            self, "new_value", _exact_nonempty_text(self.new_value, "new_value")
        )
        object.__setattr__(self, "notes", _optional_text(self.notes, "notes"))


@dataclass(frozen=True)
class ScientificReviewItem:
    """One deterministic, self-contained supervisor review row."""

    review_id: str
    data_release_id: str
    id_contaminant: str
    id_name: str
    workbook: str
    sheet: str
    cell: str
    source_row: int
    canonical_field: str
    existing_value: str
    review_type: ScientificReviewType
    finding_code: str
    finding_message: str
    proposed_value: str | None
    source_system: str | None
    source_record_id: str | None
    evidence_url: str | None
    retrieval_date: str | None
    rationale: str | None
    status: ScientificReviewStatus
    reviewer: str | None
    review_date: str | None
    notes: str | None
    implementation_status: str
    revalidation_result: str


@dataclass(frozen=True)
class ScientificReviewPackage:
    """Complete 3.3a handoff data for one inspected release."""

    data_release_id: str
    schema_version: str
    findings: tuple[ValidationFinding, ...]
    review_items: tuple[ScientificReviewItem, ...]
    resolved_cleanup: tuple[ResolvedCleanupRecord, ...]


@dataclass(frozen=True)
class ScientificReviewReconciliation:
    """One later-validation outcome for an earlier review row."""

    review_id: str
    status: ReviewReconciliationStatus
    matching_finding_code: str | None


def _glossary_columns(validated: ValidatedIdentityRelationships) -> Mapping[str, int]:
    snapshot = validated.workbook_contract.raw_pair.glossary_snapshot
    sheet = next(sheet for sheet in snapshot.sheets if sheet.name == GLOSSARY_SHEET_NAME)
    cells = {coordinate_to_tuple(cell.coordinate): cell for cell in sheet.cells}
    table = next(table for table in sheet.tables if table.name == GLOSSARY_TABLE_NAME)
    min_column, header_row, max_column, _ = range_boundaries(table.reference)
    columns_by_header = {
        cells[(header_row, column)].value: column
        for column in range(min_column, max_column + 1)
    }
    return MappingProxyType(
        {
            canonical: columns_by_header[header]
            for header, canonical in GLOSSARY_HEADER_MAP.items()
        }
    )


def _finding_by_key(
    inspection: ScientificFieldInspection,
) -> Mapping[tuple[str, str, int], tuple[ValidationFinding, ...]]:
    grouped: dict[tuple[str, str, int], list[ValidationFinding]] = {}
    for finding in inspection.findings:
        if (
            finding.id_contaminant is None
            or finding.canonical_field is None
            or finding.source_row is None
        ):
            continue
        grouped.setdefault(
            (finding.id_contaminant, finding.canonical_field, finding.source_row), []
        ).append(finding)
    return MappingProxyType(
        {key: tuple(values) for key, values in grouped.items()}
    )


def _review_id(
    release_id: str,
    id_contaminant: str,
    canonical_field: str,
    cell: str,
) -> str:
    return f"3.3a-{release_id}-{id_contaminant}-{canonical_field}-{cell}"


def _apply_proposal(
    item: ScientificReviewItem,
    proposal: ScientificReviewProposal | None,
) -> ScientificReviewItem:
    if proposal is None:
        return item
    return replace(
        item,
        proposed_value=proposal.proposed_value,
        source_system=proposal.source_system,
        source_record_id=proposal.source_record_id,
        evidence_url=proposal.evidence_url,
        retrieval_date=proposal.retrieval_date,
        rationale=proposal.rationale,
        status=proposal.status,
        reviewer=proposal.reviewer,
        review_date=proposal.review_date,
        notes=proposal.notes,
    )


def build_scientific_review_package(
    identity_relationships: ValidatedIdentityRelationships,
    proposals: Iterable[ScientificReviewProposal] = (),
    resolved_cleanup: Iterable[ResolvedCleanupRecord] = (),
) -> ScientificReviewPackage:
    """Build review rows from raw snapshots and supplied evidence only."""

    inspection = inspect_scientific_fields(identity_relationships)
    proposal_rows = tuple(proposals)
    proposal_map: dict[tuple[str, str], ScientificReviewProposal] = {}
    for proposal in proposal_rows:
        if not isinstance(proposal, ScientificReviewProposal):
            raise TypeError("proposals must contain ScientificReviewProposal")
        key = (proposal.id_contaminant, proposal.canonical_field)
        if key in proposal_map:
            raise ValueError(f"duplicate proposal for {key!r}")
        proposal_map[key] = proposal

    cleanup_rows = tuple(resolved_cleanup)
    if any(not isinstance(row, ResolvedCleanupRecord) for row in cleanup_rows):
        raise TypeError("resolved_cleanup must contain ResolvedCleanupRecord")
    cleanup_rows = tuple(
        sorted(
            cleanup_rows,
            key=lambda row: (row.id_contaminant, row.canonical_field, row.cell),
        )
    )

    names = {
        row.id_contaminant: row.id_name
        for row in identity_relationships.glossary_identities
    }
    columns = _glossary_columns(identity_relationships)
    grouped_findings = _finding_by_key(inspection)
    workbook_name = (
        identity_relationships.workbook_contract.raw_pair.glossary_snapshot.path.name
    )
    items: list[ScientificReviewItem] = []
    represented_keys: set[tuple[str, str]] = set()

    for record in inspection.contaminants:
        for canonical_field in (
            "id_casrn",
            "id_inchikey",
            "id_chem_formula",
            "source_notes_sources",
        ):
            raw_value = record.raw_values[canonical_field]
            key = (record.id_contaminant, canonical_field, record.source_row)
            findings = grouped_findings.get(key, ())
            error = next(
                (
                    finding
                    for finding in findings
                    if finding.severity is ValidationSeverity.ERROR
                ),
                None,
            )
            warning = next(
                (
                    finding
                    for finding in findings
                    if finding.severity is ValidationSeverity.WARNING
                ),
                None,
            )

            review_type: ScientificReviewType | None = None
            status: ScientificReviewStatus | None = None
            finding: ValidationFinding | None = None
            if canonical_field == "source_notes_sources" and raw_value in {None, ""}:
                review_type = ScientificReviewType.PENDING_SOURCE
                status = ScientificReviewStatus.UNKNOWN_PENDING
                finding = warning
            elif canonical_field in _IDENTIFIER_FIELDS and error is not None:
                review_type = ScientificReviewType.INVALID_SCIENTIFIC_VALUE
                status = ScientificReviewStatus.CORRECTION_REQUIRED
                finding = error
            elif canonical_field in _IDENTIFIER_FIELDS and raw_value in _NOT_APPLICABLE_TOKENS:
                review_type = ScientificReviewType.UNVERIFIED_NOT_APPLICABLE
                status = ScientificReviewStatus.NEEDS_REVIEW
                finding = warning
            elif canonical_field in _IDENTIFIER_FIELDS and raw_value is None:
                review_type = ScientificReviewType.UNKNOWN_IDENTIFIER
                status = ScientificReviewStatus.UNKNOWN_PENDING
                finding = warning
            elif canonical_field in _SCIENTIFIC_REVIEW_FIELDS and error is not None:
                review_type = ScientificReviewType.INVALID_SCIENTIFIC_VALUE
                status = ScientificReviewStatus.CORRECTION_REQUIRED
                finding = error

            if review_type is None or status is None:
                continue
            if finding is None:
                raise ValueError(
                    f"reviewable field {record.id_contaminant} {canonical_field} "
                    "has no contextual finding"
                )
            cell = f"{get_column_letter(columns[canonical_field])}{record.source_row}"
            item = ScientificReviewItem(
                review_id=_review_id(
                    inspection.data_release_id,
                    record.id_contaminant,
                    canonical_field,
                    cell,
                ),
                data_release_id=inspection.data_release_id,
                id_contaminant=record.id_contaminant,
                id_name=names[record.id_contaminant],
                workbook=workbook_name,
                sheet=GLOSSARY_SHEET_NAME,
                cell=cell,
                source_row=record.source_row,
                canonical_field=canonical_field,
                existing_value=repr(raw_value),
                review_type=review_type,
                finding_code=finding.code,
                finding_message=finding.message,
                proposed_value=None,
                source_system=None,
                source_record_id=None,
                evidence_url=None,
                retrieval_date=None,
                rationale=None,
                status=status,
                reviewer=None,
                review_date=None,
                notes=None,
                implementation_status="not_applied",
                revalidation_result="not_run",
            )
            proposal_key = (record.id_contaminant, canonical_field)
            item = _apply_proposal(item, proposal_map.get(proposal_key))
            items.append(item)
            represented_keys.add(proposal_key)

    unused = tuple(sorted(set(proposal_map) - represented_keys))
    if unused:
        raise ValueError(f"proposal does not match a review item: {unused[0]!r}")

    items.sort(
        key=lambda item: (
            item.review_type.value,
            item.id_contaminant,
            item.canonical_field,
            item.cell,
        )
    )
    return ScientificReviewPackage(
        data_release_id=inspection.data_release_id,
        schema_version=inspection.schema_version,
        findings=inspection.findings,
        review_items=tuple(items),
        resolved_cleanup=cleanup_rows,
    )


def reconcile_scientific_review_items(
    review_items: Iterable[ScientificReviewItem],
    later_findings: Iterable[ValidationFinding],
) -> tuple[ScientificReviewReconciliation, ...]:
    """Compare review rows with later findings without editing source data."""

    items = tuple(review_items)
    if any(not isinstance(item, ScientificReviewItem) for item in items):
        raise TypeError("review_items must contain ScientificReviewItem")
    findings = tuple(later_findings)
    if any(not isinstance(finding, ValidationFinding) for finding in findings):
        raise TypeError("later_findings must contain ValidationFinding")

    reconciled = []
    for item in items:
        exact = next(
            (
                finding
                for finding in findings
                if finding.id_contaminant == item.id_contaminant
                and finding.canonical_field == item.canonical_field
                and finding.source_row == item.source_row
            ),
            None,
        )
        if exact is not None:
            status = ReviewReconciliationStatus.STILL_FAILING
            code = exact.code
        else:
            replacement = next(
                (
                    finding
                    for finding in findings
                    if finding.id_contaminant == item.id_contaminant
                    and finding.canonical_field == item.canonical_field
                ),
                None,
            )
            if replacement is None:
                status = ReviewReconciliationStatus.RESOLVED
                code = None
            else:
                status = ReviewReconciliationStatus.SUPERSEDED
                code = replacement.code
        reconciled.append(
            ScientificReviewReconciliation(
                review_id=item.review_id,
                status=status,
                matching_finding_code=code,
            )
        )
    return tuple(sorted(reconciled, key=lambda row: row.review_id))


_REVIEW_HEADERS = (
    "review_id",
    "data_release_id",
    "id_contaminant",
    "id_name",
    "workbook",
    "sheet",
    "cell",
    "source_row",
    "canonical_field",
    "existing_value",
    "review_type",
    "finding_code",
    "finding_message",
    "proposed_value",
    "source_system",
    "source_record_id",
    "evidence_url",
    "retrieval_date",
    "rationale",
    "status",
    "reviewer",
    "review_date",
    "notes",
    "implementation_status",
    "revalidation_result",
)
_CLEANUP_HEADERS = (
    "id_contaminant",
    "id_name",
    "workbook",
    "sheet",
    "cell",
    "canonical_field",
    "old_value",
    "new_value",
    "correction_type",
    "notes",
)


def _excel_value(value: object) -> object:
    if isinstance(value, StrEnum):
        value = value.value
    if isinstance(value, str) and value.startswith("="):
        return f"'{value}"
    return value


def _review_row(item: ScientificReviewItem) -> list[object]:
    return [_excel_value(getattr(item, header)) for header in _REVIEW_HEADERS]


def _cleanup_row(item: ResolvedCleanupRecord) -> list[object]:
    return [_excel_value(getattr(item, header)) for header in _CLEANUP_HEADERS]


def _write_instructions(sheet, package: ScientificReviewPackage) -> None:
    rows = (
        ("Phase 3.3a Supervisor Review", None),
        ("Data release", package.data_release_id),
        ("Workbook schema", package.schema_version),
        ("Purpose", "Review proposed scientific corrections and pending values without editing authoritative workbooks."),
        ("Reviewer action", "Complete status, proposed/approved value, evidence, reviewer, date, rationale, and notes on the review sheets."),
        ("Column layout", "Generated source/finding columns come first; reviewer decision and evidence columns follow with explicit names."),
        ("Important", "A proposal is not authoritative until it is approved, applied on a data-only branch, assigned a new workbook revision, and revalidated."),
        ("Blank optional source", "Unknown/pending. It remains visible as a warning and canonical null."),
        ("Blank identifier", "Unknown/pending. Do not convert it to not applicable without review."),
        ("NA/N/A identifier", "Permitted representation where the schema allows it, but it still needs a per-ID rationale."),
        ("Allowed statuses", ", ".join(_REVIEW_STATUSES)),
        ("Authoritative workbook edits", "None are performed by this review workbook generator."),
    )
    for row in rows:
        sheet.append([_excel_value(value) for value in row])


def write_supervisor_review_workbook(
    package: ScientificReviewPackage,
    output_path: Path | None = None,
) -> Path:
    """Atomically write the separate review-only XLSX workbook."""

    if not isinstance(package, ScientificReviewPackage):
        raise TypeError("package must be ScientificReviewPackage")
    path = (
        supervisor_review_workbook_path(package.data_release_id)
        if output_path is None
        else Path(output_path)
    )
    if path.suffix.lower() != ".xlsx":
        raise ValueError("supervisor review output must use .xlsx")
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.{uuid4().hex}.tmp.xlsx")

    workbook = Workbook()
    workbook.properties.creator = "Red Hill contaminant pipeline"
    workbook.properties.created = datetime(2000, 1, 1)
    workbook.properties.modified = datetime(2000, 1, 1)
    instructions = workbook.active
    instructions.title = "Instructions"
    _write_instructions(instructions, package)

    identifier_sheet = workbook.create_sheet("Identifier Review")
    identifier_sheet.append(list(_REVIEW_HEADERS))
    pending_sheet = workbook.create_sheet("Pending Sources")
    pending_sheet.append(list(_REVIEW_HEADERS))
    cleanup_sheet = workbook.create_sheet("Resolved Cleanup")
    cleanup_sheet.append(list(_CLEANUP_HEADERS))

    for item in package.review_items:
        target = (
            pending_sheet
            if item.review_type is ScientificReviewType.PENDING_SOURCE
            else identifier_sheet
        )
        target.append(_review_row(item))
    for item in package.resolved_cleanup:
        cleanup_sheet.append(_cleanup_row(item))

    try:
        workbook.save(temporary)
        temporary.replace(path)
    finally:
        workbook.close()
        if temporary.exists():
            temporary.unlink()
    return path
